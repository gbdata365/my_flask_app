# -*- coding: utf-8 -*-
"""
================================================================================
인구 집계표 대시보드 (edu_dash2.py)
================================================================================

[목적]
- PDF 형식의 인구 집계표를 웹에서 조회하고 내보내기
- 연령별, 시도별, 청년/노령 인구 현황 표시
- 다양한 형식(Excel, MD, HTML)으로 내보내기 지원

[주요 기능]
1. 연령별 인구 현황 (5세별/10세별/정책연령 - 코드테이블 기반)
2. 시도별 인구 현황
3. 청년인구 (코드테이블의 청년 정의 사용)
4. 노령인구 (65세 이상)
5. 고령화율 추이
6. 필터: 권역, 시도, 시군구, 기준년월, 연령구분

[코드 테이블]
- code_age_group: 연령그룹 정의 (category별로 구분)
  - category 1: 5세별 (21개 그룹)
  - category 2: 10세별 (15개 그룹)
  - category 3: 정책연령 (15개 그룹)

[기술 스택]
- Backend: Flask (Python)
- Frontend: Tailwind CSS + JavaScript
- Database: PostgreSQL
- Export: openpyxl (Excel), Markdown, HTML

================================================================================
"""
import sys
from pathlib import Path
import json
import pandas as pd
import numpy as np
import io
import base64
from datetime import datetime
import koreanize_matplotlib
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from jinja2 import Environment, FileSystemLoader

sys.path.insert(0, str(Path(__file__).parent.parent.parent))
from module.db import get_db_engine
from module.menu_generator import MenuGenerator
from flask import jsonify, Response, send_file, request

# ================================================================================
# [모듈화] 공통 내보내기 유틸리티 - 다른 분야에서도 재사용 가능
# 정렬 파라미터 지원, 탭 형식 HTML, 차트 포함 내보내기
# ================================================================================
from common.export_utils import DataExporter

POP_BASE = Path(__file__).parent.parent
TEMPLATE_DIR = POP_BASE / 'templates'

# 차트 설정
try:
    from config.chart_config import (
        POPULATION_UNIT, SINGLE_AGE_UNIT, REGIONAL_SUBPLOT, ACCORDION_TABLE,
        get_unit_config, convert_value
    )
except ImportError:
    # 설정 파일 없으면 기본값
    POPULATION_UNIT = {'unit': 100, 'label': '백 명', 'format': '{:,.0f}'}
    SINGLE_AGE_UNIT = {'unit': 100, 'label': '백 가구', 'format': '{:,.0f}'}
    REGIONAL_SUBPLOT = {'max_regions': None, 'cols': 3, 'fig_width_per_col': 5, 'fig_height_per_row': 4}
    ACCORDION_TABLE = {'expand_all': True}

# Jinja2 환경 설정
jinja_env = Environment(loader=FileSystemLoader(str(TEMPLATE_DIR)))


# =============================================================================
# 코드 테이블 조회 함수
# =============================================================================

def get_age_groups_from_db(category=1):
    """
    code_age_group 테이블에서 연령그룹 정보 조회

    Args:
        category (int): 연령 카테고리
            1 = 5세별 (21개)
            2 = 10세별 (15개)
            3 = 정책연령 (15개)

    Returns:
        DataFrame: code, code_name, column_name, age_start, age_end, sort_order
    """
    engine = get_db_engine()
    df = pd.read_sql(f"""
        SELECT code, code_name, column_name, age_start, age_end, sort_order
        FROM code_age_group
        WHERE category = {category}
          AND is_active = TRUE
        ORDER BY sort_order
    """, engine)
    return df


def get_age_categories():
    """
    연령 카테고리 목록 조회

    Returns:
        list: [{'category': 1, 'category_name': '5세별'}, ...]
    """
    engine = get_db_engine()
    df = pd.read_sql("""
        SELECT DISTINCT category, category_name
        FROM code_age_group
        WHERE is_active = TRUE
        ORDER BY category
    """, engine)
    return df.to_dict('records')


def get_youth_age_groups():
    """
    청년인구 정의 조회 (정책연령 카테고리에서)

    Returns:
        dict: {'19_39': [...columns...], '19_34': [...columns...]}
    """
    engine = get_db_engine()

    # 정책연령(category=3)에서 청년 관련 그룹 조회
    df = pd.read_sql("""
        SELECT code_name, column_name, age_start, age_end
        FROM code_age_group
        WHERE category = 3
          AND is_active = TRUE
          AND (code_name LIKE '%청년%' OR (age_start >= 19 AND age_end <= 39))
        ORDER BY sort_order
    """, engine)

    # 5세별(category=1)에서 청년 연령대 추출
    df_5 = pd.read_sql("""
        SELECT column_name, age_start, age_end
        FROM code_age_group
        WHERE category = 1
          AND is_active = TRUE
          AND age_start >= 15 AND age_end <= 39
        ORDER BY sort_order
    """, engine)

    return {
        'youth_columns': df_5.to_dict('records'),
        'policy_youth': df.to_dict('records') if len(df) > 0 else None
    }


def get_elderly_definition():
    """
    노령인구 정의 조회

    Returns:
        dict: {'column': 'elderly_pop', 'age_start': 65, 'description': '65세 이상'}
    """
    # cache_sigungu_indicators에 elderly_pop 컬럼이 있음
    return {
        'column': 'elderly_pop',
        'age_start': 65,
        'description': '65세 이상'
    }


def get_active_indicators(category=1):
    """
    code_indicator 테이블에서 is_active=true인 지표 조회

    Args:
        category (int): 지표 카테고리 (1=인구지표, 2=가구지표)

    Returns:
        list: [{'column_name': ..., 'display_name': ..., 'description': ...}, ...]
    """
    engine = get_db_engine()
    df = pd.read_sql(f"""
        SELECT column_name, display_name, description, numerator, denominator, multiplier, decimal_places
        FROM code_indicator
        WHERE is_active = true AND category = {category}
        ORDER BY sort_order
    """, engine)
    return df.to_dict('records')


def get_column_name_labels():
    """
    DB 코드 테이블에서 컬럼명 → 한글 라벨 매핑 조회

    Returns:
        dict: {'youth_pop': '유소년', 'elderly_pop': '고령인구', ...}
    """
    engine = get_db_engine()
    labels = {}

    # code_age_group에서 column_name → code_name 매핑
    try:
        age_df = pd.read_sql("""
            SELECT column_name, code_name
            FROM code_age_group
            WHERE is_active = TRUE AND column_name IS NOT NULL AND column_name != ''
        """, engine)
        for _, row in age_df.iterrows():
            if row['column_name']:
                labels[row['column_name']] = row['code_name']
    except:
        pass

    # code_indicator에서 numerator/denominator용 매핑 추가
    try:
        ind_df = pd.read_sql("""
            SELECT column_name, display_name, numerator, denominator
            FROM code_indicator
            WHERE is_active = TRUE
        """, engine)
        for _, row in ind_df.iterrows():
            if row['column_name']:
                labels[row['column_name']] = row['display_name']
    except:
        pass

    # 기본 매핑 추가 (DB에 없는 경우 대비)
    default_labels = {
        'total_pop': '총인구',
        'male_pop': '남자인구',
        'female_pop': '여자인구',
        'household_cnt': '전체가구',
        'single_cnt': '1인가구',
        'couple_cnt': '부부가구',
        'youth_pop': '유소년인구',
        'working_pop': '생산가능인구',
        'elderly_pop': '노령인구',
        'late_elderly_pop': '후기고령인구',
        'early_elderly_pop': '전기고령인구',
        'youth_pop_19_39': '청년인구(19-39)',
        'youth_pop_19_34': '청년인구(19-34)'
    }

    # DB에서 가져온 값이 없는 경우만 기본값 사용
    for k, v in default_labels.items():
        if k not in labels:
            labels[k] = v

    return labels


# =============================================================================
# 필터 옵션 조회
# =============================================================================

def get_filter_options():
    """필터 옵션 조회 (코드테이블 포함)"""
    engine = get_db_engine()

    # 기준년월 목록
    ym_df = pd.read_sql("""
        SELECT DISTINCT TO_CHAR(base_ym, 'YYYYMM') as ym
        FROM cache_sigungu_indicators
        ORDER BY ym DESC
    """, engine)

    # 시도 목록 (강원/전북 특별자치도 통합 - 최신 명칭으로)
    sido_df = pd.read_sql("""
        SELECT DISTINCT
            CASE
                WHEN sido_nm IN ('강원특별자치도', '강원도') THEN '강원특별자치도'
                WHEN sido_nm IN ('전북특별자치도', '전라북도') THEN '전북특별자치도'
                WHEN sido_nm IN ('제주특별자치도', '제주도') THEN '제주특별자치도'
                ELSE sido_nm
            END as sido_nm,
            MIN(LEFT(sigungu_code, 2)) as sido_code
        FROM dim_admin_area
        WHERE sido_nm IS NOT NULL
        GROUP BY CASE
            WHEN sido_nm IN ('강원특별자치도', '강원도') THEN '강원특별자치도'
            WHEN sido_nm IN ('전북특별자치도', '전라북도') THEN '전북특별자치도'
            WHEN sido_nm IN ('제주특별자치도', '제주도') THEN '제주특별자치도'
            ELSE sido_nm
        END
        ORDER BY sido_code
    """, engine)

    # 권역 목록
    region_df = pd.read_sql("""
        SELECT DISTINCT region_nm
        FROM dim_admin_area
        WHERE region_nm IS NOT NULL
        ORDER BY region_nm
    """, engine)

    # 연령 카테고리 목록
    age_categories = get_age_categories()

    # 활성화된 인구지표 목록 (동적 탭용)
    active_indicators = get_active_indicators(1)

    return {
        'base_ym_list': ym_df['ym'].tolist(),
        'sido_list': sido_df['sido_nm'].tolist(),
        'region_list': region_df['region_nm'].tolist(),
        'age_categories': age_categories,
        'active_indicators': active_indicators,
    }


# =============================================================================
# 데이터 조회 함수들
# =============================================================================

def get_age_population_by_region(base_ym_list, age_category=1, aggregate_type='sido', region=None, priority_region='경상북도'):
    """
    지역별 연령 인구 데이터 조회 (subplot 차트용)

    Args:
        base_ym_list: 기준년월 리스트
        age_category: 연령 카테고리 (1=5세별, 2=10세별, 3=정책연령)
        aggregate_type: 'region' (권역별) 또는 'sido' (시도별)
        region: 특정 권역 선택 시
        priority_region: 우선 표시할 지역명 (기본: 경상북도)

    Returns:
        dict: {지역명: {'total_pop': {ym: pop}, 'age_data': [{age_group: ..., pop_YYYYMM: ...}, ...]}, ...}
    """
    engine = get_db_engine()

    # 코드테이블에서 연령그룹 조회
    age_groups = get_age_groups_from_db(age_category)
    if age_groups.empty:
        return {}

    age_columns = [(row['column_name'], row['code_name']) for _, row in age_groups.iterrows()]
    cols_sum = ', '.join([f'SUM(COALESCE({col}, 0)) as {col}' for col, _ in age_columns])

    # 지역별 그룹화
    if aggregate_type == 'region':
        group_col = "d.region_nm"
        group_name = "region_nm"
        join_clause = """
            JOIN (
                SELECT DISTINCT sigungu_code, region_nm, region_code
                FROM dim_admin_area
                WHERE region_nm IS NOT NULL
            ) d ON c.sigungu_code = d.sigungu_code
        """
        where_clause = ""
        order_clause = "ORDER BY MIN(d.region_code)"
    else:
        # 시도별
        group_col = """CASE
            WHEN c.sido_nm IN ('강원특별자치도', '강원도') THEN '강원특별자치도'
            WHEN c.sido_nm IN ('전북특별자치도', '전라북도') THEN '전북특별자치도'
            WHEN c.sido_nm IN ('제주특별자치도', '제주도') THEN '제주특별자치도'
            ELSE c.sido_nm
        END"""
        group_name = "sido_nm"
        join_clause = """
            JOIN (
                SELECT DISTINCT sigungu_code, region_nm
                FROM dim_admin_area
                WHERE region_nm IS NOT NULL
            ) d ON c.sigungu_code = d.sigungu_code
        """
        where_clause = f"AND d.region_nm = '{region}'" if region else ""
        order_clause = "ORDER BY MIN(LEFT(c.sigungu_code, 2))"

    regional_data = {}

    for ym in base_ym_list:
        df = pd.read_sql(f"""
            SELECT
                {group_col} as {group_name},
                SUM(total_pop) as total_pop,
                {cols_sum}
            FROM cache_sigungu_indicators c
            {join_clause}
            WHERE TO_CHAR(c.base_ym, 'YYYYMM') = '{ym}'
            {where_clause}
            GROUP BY {group_col}
            {order_clause}
        """, engine)

        for _, row in df.iterrows():
            region_name = row[group_name]
            if region_name not in regional_data:
                regional_data[region_name] = {
                    'total_pop': {},
                    'age_dict': {col_name: {'age_group': label} for col_name, label in age_columns}
                }

            # 총인구 저장
            regional_data[region_name]['total_pop'][ym] = int(row['total_pop'])

            for col_name, label in age_columns:
                regional_data[region_name]['age_dict'][col_name][f'pop_{ym}'] = int(row[col_name])

    # 딕셔너리를 정렬된 형태로 변환 (priority_region을 맨 앞으로)
    result = {}

    # 우선 지역이 있으면 먼저 추가
    if priority_region and priority_region in regional_data:
        data = regional_data[priority_region]
        result[priority_region] = {
            'total_pop': data['total_pop'],
            'age_data': list(data['age_dict'].values())
        }

    # 나머지 지역 추가
    for region_name, data in regional_data.items():
        if region_name != priority_region:
            result[region_name] = {
                'total_pop': data['total_pop'],
                'age_data': list(data['age_dict'].values())
            }

    return result


def get_age_population_table(base_ym_list, age_category=1):
    """
    연령별 인구 현황 테이블 (코드테이블 기반)

    Args:
        base_ym_list: 기준년월 리스트
        age_category: 연령 카테고리 (1=5세별, 2=10세별, 3=정책연령)
    """
    engine = get_db_engine()

    # 코드테이블에서 연령그룹 조회
    age_groups = get_age_groups_from_db(age_category)

    if age_groups.empty:
        return {'headers': [], 'data': [], 'category_name': ''}

    # 카테고리명 조회
    cat_name_df = pd.read_sql(f"""
        SELECT DISTINCT category_name
        FROM code_age_group
        WHERE category = {age_category}
        LIMIT 1
    """, engine)
    category_name = cat_name_df['category_name'].values[0] if len(cat_name_df) > 0 else ''

    # 연령그룹 컬럼과 라벨 추출
    age_columns = [(row['column_name'], row['code_name']) for _, row in age_groups.iterrows()]

    results = []
    for ym in base_ym_list:
        cols_sum = ', '.join([f'SUM(COALESCE({col}, 0)) as {col}' for col, _ in age_columns])
        df = pd.read_sql(f"""
            SELECT
                '{ym}' as base_ym,
                SUM(total_pop) as total_pop,
                {cols_sum}
            FROM cache_sigungu_indicators
            WHERE TO_CHAR(base_ym, 'YYYYMM') = '{ym}'
        """, engine)
        results.append(df)

    if not results:
        return {'headers': [], 'data': [], 'category_name': category_name}

    combined_df = pd.concat(results, ignore_index=True)

    # 결과 포맷팅
    data = []

    # 총인구 행 먼저
    total_row = {'age_group': '총 인구수'}
    for i, ym in enumerate(base_ym_list):
        total_row[f'pop_{ym}'] = int(combined_df[combined_df['base_ym'] == ym]['total_pop'].values[0])
        # 가장 오래된 시점(마지막)에는 증감률 표시 안함 - 직전 시점과 비교
        if i < len(base_ym_list) - 1:
            next_ym = base_ym_list[i+1]  # 직전(더 오래된) 시점
            next_pop = int(combined_df[combined_df['base_ym'] == next_ym]['total_pop'].values[0])
            curr_pop = total_row[f'pop_{ym}']
            total_row[f'rate_{ym}'] = round((curr_pop - next_pop) / next_pop * 100, 2) if next_pop > 0 else 0
    data.append(total_row)

    # 각 연령그룹 행
    for col_name, label in age_columns:
        row = {'age_group': label}
        for i, ym in enumerate(base_ym_list):
            try:
                pop = int(combined_df[combined_df['base_ym'] == ym][col_name].values[0])
            except:
                pop = 0
            row[f'pop_{ym}'] = pop

            # 가장 오래된 시점(마지막)에는 증감률 표시 안함 - 직전 시점과 비교
            if i < len(base_ym_list) - 1:
                next_ym = base_ym_list[i+1]  # 직전(더 오래된) 시점
                try:
                    next_pop = int(combined_df[combined_df['base_ym'] == next_ym][col_name].values[0])
                except:
                    next_pop = 0
                row[f'rate_{ym}'] = round((pop - next_pop) / next_pop * 100, 2) if next_pop > 0 else 0
        data.append(row)

    return {
        'headers': base_ym_list,
        'data': data,
        'category_name': category_name
    }


def get_region_population_table(base_ym_list):
    """
    권역별 인구 현황 테이블 (5개 권역 + 3개 특별자치도)
    - dim_admin_area에서 시군구별 region_nm을 DISTINCT로 조회하여 중복 조인 방지
    """
    engine = get_db_engine()

    results = []
    for ym in base_ym_list:
        df = pd.read_sql(f"""
            SELECT
                d.region_nm,
                d.region_code,
                SUM(c.total_pop) as total_pop,
                SUM(COALESCE(c.household_cnt, 0)) as household_cnt,
                SUM(COALESCE(c.elderly_pop, 0)) as elderly_pop,
                ROUND(SUM(COALESCE(c.elderly_pop, 0))::numeric / NULLIF(SUM(c.total_pop), 0) * 100, 2) as elderly_ratio
            FROM cache_sigungu_indicators c
            JOIN (
                SELECT DISTINCT sigungu_code, region_nm, region_code
                FROM dim_admin_area
                WHERE region_nm IS NOT NULL
            ) d ON c.sigungu_code = d.sigungu_code
            WHERE TO_CHAR(c.base_ym, 'YYYYMM') = '{ym}'
            GROUP BY d.region_nm, d.region_code
            ORDER BY d.region_code
        """, engine)
        df['base_ym'] = ym
        results.append(df)

    if not results:
        return {'headers': [], 'data': []}

    combined_df = pd.concat(results, ignore_index=True)

    # region_code 순서로 정렬된 권역 목록 생성
    region_order_df = combined_df[['region_nm', 'region_code']].drop_duplicates().sort_values('region_code')
    region_order = region_order_df['region_nm'].tolist()

    pivot_pop = combined_df.pivot(index='region_nm', columns='base_ym', values='total_pop').fillna(0)
    pivot_household = combined_df.pivot(index='region_nm', columns='base_ym', values='household_cnt').fillna(0)
    pivot_elderly_ratio = combined_df.pivot(index='region_nm', columns='base_ym', values='elderly_ratio').fillna(0)

    # 권역코드 순서로 정렬
    pivot_pop = pivot_pop.reindex([r for r in region_order if r in pivot_pop.index])
    pivot_household = pivot_household.reindex([r for r in region_order if r in pivot_household.index])
    pivot_elderly_ratio = pivot_elderly_ratio.reindex([r for r in region_order if r in pivot_elderly_ratio.index])

    data = []

    # 전국 합계 먼저
    total_row = {'name': '전국'}
    for i, ym in enumerate(base_ym_list):
        if ym in pivot_pop.columns:
            total_row[f'pop_{ym}'] = int(pivot_pop[ym].sum())
            total_row[f'household_{ym}'] = int(pivot_household[ym].sum())
            # 가장 오래된 시점(마지막)에는 증감률 표시 안함 - 직전 시점과 비교
            if i < len(base_ym_list) - 1 and base_ym_list[i+1] in pivot_pop.columns:
                next_pop = int(pivot_pop[base_ym_list[i+1]].sum())
                total_row[f'pop_rate_{ym}'] = round((total_row[f'pop_{ym}'] - next_pop) / next_pop * 100, 2) if next_pop > 0 else 0
    data.append(total_row)

    # 각 권역 행
    for region in pivot_pop.index:
        row = {'name': region}
        for i, ym in enumerate(base_ym_list):
            if ym in pivot_pop.columns:
                pop = int(pivot_pop.loc[region, ym]) if pd.notna(pivot_pop.loc[region, ym]) else 0
                household = int(pivot_household.loc[region, ym]) if pd.notna(pivot_household.loc[region, ym]) else 0
                row[f'pop_{ym}'] = pop
                row[f'household_{ym}'] = household

                # 가장 오래된 시점(마지막)에는 증감률 표시 안함 - 직전 시점과 비교
                if i < len(base_ym_list) - 1 and base_ym_list[i+1] in pivot_pop.columns:
                    next_pop = int(pivot_pop.loc[region, base_ym_list[i+1]]) if pd.notna(pivot_pop.loc[region, base_ym_list[i+1]]) else 0
                    row[f'pop_rate_{ym}'] = round((pop - next_pop) / next_pop * 100, 2) if next_pop > 0 else 0
        data.append(row)

    return {
        'headers': base_ym_list,
        'data': data
    }


def get_sido_population_table(base_ym_list, region=None):
    """시도별 인구 현황 테이블"""
    engine = get_db_engine()

    where_clause = ""
    if region:
        where_clause = f"AND region_nm = '{region}'"

    results = []
    for ym in base_ym_list:
        df = pd.read_sql(f"""
            SELECT
                CASE
                    WHEN sido_nm IN ('강원특별자치도', '강원도') THEN '강원특별자치도'
                    WHEN sido_nm IN ('전북특별자치도', '전라북도') THEN '전북특별자치도'
                    WHEN sido_nm IN ('제주특별자치도', '제주도') THEN '제주특별자치도'
                    ELSE sido_nm
                END as sido_nm,
                MIN(LEFT(sigungu_code, 2)) as sido_code,
                SUM(total_pop) as total_pop,
                SUM(COALESCE(household_cnt, 0)) as household_cnt
            FROM cache_sigungu_indicators
            WHERE TO_CHAR(base_ym, 'YYYYMM') = '{ym}'
            {where_clause}
            GROUP BY CASE
                    WHEN sido_nm IN ('강원특별자치도', '강원도') THEN '강원특별자치도'
                    WHEN sido_nm IN ('전북특별자치도', '전라북도') THEN '전북특별자치도'
                    WHEN sido_nm IN ('제주특별자치도', '제주도') THEN '제주특별자치도'
                    ELSE sido_nm
                END
            ORDER BY MIN(LEFT(sigungu_code, 2))
        """, engine)
        df['base_ym'] = ym
        results.append(df)

    if not results:
        return {'headers': [], 'data': []}

    combined_df = pd.concat(results, ignore_index=True)

    # 시도코드로 그룹화하여 정렬 순서 유지
    sido_order = combined_df.groupby('sido_nm')['sido_code'].first().sort_values()

    pivot_pop = combined_df.pivot(index='sido_nm', columns='base_ym', values='total_pop').fillna(0)
    pivot_household = combined_df.pivot(index='sido_nm', columns='base_ym', values='household_cnt').fillna(0)

    # 시도코드 순으로 정렬
    pivot_pop = pivot_pop.reindex(sido_order.index)
    pivot_household = pivot_household.reindex(sido_order.index)

    data = []

    # 전국 합계 먼저
    total_row = {'name': '전국'}
    for i, ym in enumerate(base_ym_list):
        if ym in pivot_pop.columns:
            total_row[f'pop_{ym}'] = int(pivot_pop[ym].sum())
            total_row[f'household_{ym}'] = int(pivot_household[ym].sum())
            # 가장 오래된 시점(마지막)에는 증감률 표시 안함 - 직전 시점과 비교
            if i < len(base_ym_list) - 1 and base_ym_list[i+1] in pivot_pop.columns:
                next_pop = int(pivot_pop[base_ym_list[i+1]].sum())
                next_household = int(pivot_household[base_ym_list[i+1]].sum())
                total_row[f'pop_rate_{ym}'] = round((total_row[f'pop_{ym}'] - next_pop) / next_pop * 100, 2) if next_pop > 0 else 0
                total_row[f'household_rate_{ym}'] = round((total_row[f'household_{ym}'] - next_household) / next_household * 100, 2) if next_household > 0 else 0
    data.append(total_row)

    # 각 시도 행 (시도코드 순)
    for sido in pivot_pop.index:
        row = {'name': sido}
        for i, ym in enumerate(base_ym_list):
            if ym in pivot_pop.columns:
                pop = int(pivot_pop.loc[sido, ym]) if pd.notna(pivot_pop.loc[sido, ym]) else 0
                household = int(pivot_household.loc[sido, ym]) if pd.notna(pivot_household.loc[sido, ym]) else 0
                row[f'pop_{ym}'] = pop
                row[f'household_{ym}'] = household

                # 가장 오래된 시점(마지막)에는 증감률 표시 안함 - 직전 시점과 비교
                if i < len(base_ym_list) - 1 and base_ym_list[i+1] in pivot_pop.columns:
                    next_pop = int(pivot_pop.loc[sido, base_ym_list[i+1]]) if pd.notna(pivot_pop.loc[sido, base_ym_list[i+1]]) else 0
                    next_household = int(pivot_household.loc[sido, base_ym_list[i+1]]) if pd.notna(pivot_household.loc[sido, base_ym_list[i+1]]) else 0
                    row[f'pop_rate_{ym}'] = round((pop - next_pop) / next_pop * 100, 2) if next_pop > 0 else 0
                    row[f'household_rate_{ym}'] = round((household - next_household) / next_household * 100, 2) if next_household > 0 else 0
        data.append(row)

    return {
        'headers': base_ym_list,
        'data': data
    }


def get_youth_population_table(base_ym_list, age_range='19_39', region=None, sido=None):
    """
    청년인구 테이블 (코드테이블 기반)

    age_range:
        '19_39' - 19~39세 (청년기본법)
        '19_34' - 19~34세
    """
    engine = get_db_engine()

    where_clause = ""
    if region:
        where_clause = f"AND region_nm = '{region}'"
    if sido:
        # 특별자치도 명칭 매핑 (과거/현재 명칭 모두 검색)
        if sido == '강원특별자치도':
            where_clause += " AND sido_nm IN ('강원특별자치도', '강원도')"
        elif sido == '전북특별자치도':
            where_clause += " AND sido_nm IN ('전북특별자치도', '전라북도')"
        elif sido == '제주특별자치도':
            where_clause += " AND sido_nm IN ('제주특별자치도', '제주도')"
        else:
            where_clause += f" AND sido_nm = '{sido}'"

    # 5세별 연령그룹에서 해당 연령대 컬럼 조회
    if age_range == '19_39':
        age_end = 39
        title = '청년인구 (19~39세)'
    else:
        age_end = 34
        title = '청년인구 (19~34세)'

    # 코드테이블에서 해당 연령대 컬럼 조회
    age_cols_df = pd.read_sql(f"""
        SELECT column_name, age_start, age_end
        FROM code_age_group
        WHERE category = 1
          AND is_active = TRUE
          AND age_start >= 15 AND age_end <= {age_end + 5}
        ORDER BY sort_order
    """, engine)

    # 청년인구 계산 SQL 생성
    # 15~19세는 19세 이상만 포함 (비율 적용)
    youth_cols = []
    for _, row in age_cols_df.iterrows():
        col = row['column_name']
        start = row['age_start']
        end = row['age_end']

        if start < 19 and end >= 19:
            # 15~19세 구간: 19세 비율만 계산 (약 1/5)
            youth_cols.append(f"COALESCE({col}, 0) * 0.2")
        elif start >= 19 and end <= age_end:
            youth_cols.append(f"COALESCE({col}, 0)")
        elif start <= age_end and end > age_end:
            # 부분 포함 구간
            ratio = (age_end - start + 1) / (end - start + 1)
            youth_cols.append(f"COALESCE({col}, 0) * {ratio:.2f}")

    if not youth_cols:
        return {'title': title, 'headers': [], 'data': []}

    youth_sum_sql = ' + '.join(youth_cols)

    results = []
    for ym in base_ym_list:
        df = pd.read_sql(f"""
            SELECT
                CASE
                    WHEN sido_nm IN ('강원특별자치도', '강원도') THEN '강원특별자치도'
                    WHEN sido_nm IN ('전북특별자치도', '전라북도') THEN '전북특별자치도'
                    WHEN sido_nm IN ('제주특별자치도', '제주도') THEN '제주특별자치도'
                    ELSE sido_nm
                END as sido_nm,
                MIN(LEFT(sigungu_code, 2)) as sido_code,
                SUM({youth_sum_sql}) as youth_pop,
                SUM(total_pop) as total_pop
            FROM cache_sigungu_indicators
            WHERE TO_CHAR(base_ym, 'YYYYMM') = '{ym}'
            {where_clause}
            GROUP BY CASE
                    WHEN sido_nm IN ('강원특별자치도', '강원도') THEN '강원특별자치도'
                    WHEN sido_nm IN ('전북특별자치도', '전라북도') THEN '전북특별자치도'
                    WHEN sido_nm IN ('제주특별자치도', '제주도') THEN '제주특별자치도'
                    ELSE sido_nm
                END
            ORDER BY MIN(LEFT(sigungu_code, 2))
        """, engine)
        df['base_ym'] = ym
        results.append(df)

    if not results:
        return {'title': title, 'headers': [], 'data': []}

    combined_df = pd.concat(results, ignore_index=True)

    # 시도코드로 그룹화하여 정렬 순서 유지
    sido_order = combined_df.groupby('sido_nm')['sido_code'].first().sort_values()

    pivot = combined_df.pivot(index='sido_nm', columns='base_ym', values='youth_pop').fillna(0)

    # 시도코드 순으로 정렬
    pivot = pivot.reindex(sido_order.index)

    data = []

    # 전국 합계
    total_row = {'name': '전국'}
    for i, ym in enumerate(base_ym_list):
        if ym in pivot.columns:
            total_row[f'pop_{ym}'] = int(pivot[ym].sum())
            # 가장 오래된 시점(마지막)에는 증감률 표시 안함 - 직전 시점과 비교
            if i < len(base_ym_list) - 1 and base_ym_list[i+1] in pivot.columns:
                next_pop = int(pivot[base_ym_list[i+1]].sum())
                total_row[f'rate_{ym}'] = round((total_row[f'pop_{ym}'] - next_pop) / next_pop * 100, 2) if next_pop > 0 else 0
    data.append(total_row)

    # 각 시도 행 (시도코드 순)
    for sido in pivot.index:
        row = {'name': sido}
        for i, ym in enumerate(base_ym_list):
            if ym in pivot.columns:
                pop = int(pivot.loc[sido, ym])
                row[f'pop_{ym}'] = pop

                # 가장 오래된 시점(마지막)에는 증감률 표시 안함 - 직전 시점과 비교
                if i < len(base_ym_list) - 1 and base_ym_list[i+1] in pivot.columns:
                    next_pop = int(pivot.loc[sido, base_ym_list[i+1]])
                    row[f'rate_{ym}'] = round((pop - next_pop) / next_pop * 100, 2) if next_pop > 0 else 0
        data.append(row)

    return {
        'title': title,
        'headers': base_ym_list,
        'data': data
    }


def get_elderly_population_table(base_ym_list, region=None, sido=None):
    """
    노령인구 테이블 (65세 이상)
    - elderly_pop 컬럼 사용 (cache_sigungu_indicators)
    """
    engine = get_db_engine()

    elderly_def = get_elderly_definition()

    where_clause = ""
    if region:
        where_clause = f"AND region_nm = '{region}'"
    if sido:
        # 특별자치도 명칭 매핑 (과거/현재 명칭 모두 검색)
        if sido == '강원특별자치도':
            where_clause += " AND sido_nm IN ('강원특별자치도', '강원도')"
        elif sido == '전북특별자치도':
            where_clause += " AND sido_nm IN ('전북특별자치도', '전라북도')"
        elif sido == '제주특별자치도':
            where_clause += " AND sido_nm IN ('제주특별자치도', '제주도')"
        else:
            where_clause += f" AND sido_nm = '{sido}'"

    results = []
    for ym in base_ym_list:
        df = pd.read_sql(f"""
            SELECT
                CASE
                    WHEN sido_nm IN ('강원특별자치도', '강원도') THEN '강원특별자치도'
                    WHEN sido_nm IN ('전북특별자치도', '전라북도') THEN '전북특별자치도'
                    WHEN sido_nm IN ('제주특별자치도', '제주도') THEN '제주특별자치도'
                    ELSE sido_nm
                END as sido_nm,
                MIN(LEFT(sigungu_code, 2)) as sido_code,
                SUM(COALESCE({elderly_def['column']}, 0)) as elderly_pop,
                SUM(total_pop) as total_pop
            FROM cache_sigungu_indicators
            WHERE TO_CHAR(base_ym, 'YYYYMM') = '{ym}'
            {where_clause}
            GROUP BY CASE
                    WHEN sido_nm IN ('강원특별자치도', '강원도') THEN '강원특별자치도'
                    WHEN sido_nm IN ('전북특별자치도', '전라북도') THEN '전북특별자치도'
                    WHEN sido_nm IN ('제주특별자치도', '제주도') THEN '제주특별자치도'
                    ELSE sido_nm
                END
            ORDER BY MIN(LEFT(sigungu_code, 2))
        """, engine)
        df['base_ym'] = ym
        results.append(df)

    if not results:
        return {'headers': [], 'data': [], 'description': elderly_def['description']}

    combined_df = pd.concat(results, ignore_index=True)

    # 시도코드로 그룹화하여 정렬 순서 유지
    sido_order = combined_df.groupby('sido_nm')['sido_code'].first().sort_values()

    pivot_elderly = combined_df.pivot(index='sido_nm', columns='base_ym', values='elderly_pop').fillna(0)
    pivot_total = combined_df.pivot(index='sido_nm', columns='base_ym', values='total_pop').fillna(0)

    # 시도코드 순으로 정렬
    pivot_elderly = pivot_elderly.reindex(sido_order.index)
    pivot_total = pivot_total.reindex(sido_order.index)

    data = []

    # 전국 합계
    total_row = {'name': '전국'}
    for i, ym in enumerate(base_ym_list):
        if ym in pivot_elderly.columns:
            total_elderly = int(pivot_elderly[ym].sum())
            total_pop = int(pivot_total[ym].sum())
            total_row[f'pop_{ym}'] = total_elderly
            total_row[f'aging_rate_{ym}'] = round(total_elderly / total_pop * 100, 2) if total_pop > 0 else 0
            # 가장 오래된 시점(마지막)에는 증감률 표시 안함 - 직전 시점과 비교
            if i < len(base_ym_list) - 1 and base_ym_list[i+1] in pivot_elderly.columns:
                next_elderly = int(pivot_elderly[base_ym_list[i+1]].sum())
                total_row[f'change_rate_{ym}'] = round((total_elderly - next_elderly) / next_elderly * 100, 2) if next_elderly > 0 else 0
    data.append(total_row)

    # 각 시도 행 (시도코드 순)
    for sido in pivot_elderly.index:
        row = {'name': sido}
        for i, ym in enumerate(base_ym_list):
            if ym in pivot_elderly.columns:
                elderly = int(pivot_elderly.loc[sido, ym]) if pd.notna(pivot_elderly.loc[sido, ym]) else 0
                total = int(pivot_total.loc[sido, ym]) if pd.notna(pivot_total.loc[sido, ym]) else 0
                rate = round(elderly / total * 100, 2) if total > 0 else 0

                row[f'pop_{ym}'] = elderly
                row[f'aging_rate_{ym}'] = rate

                # 가장 오래된 시점(마지막)에는 증감률 표시 안함 - 직전 시점과 비교
                if i < len(base_ym_list) - 1 and base_ym_list[i+1] in pivot_elderly.columns:
                    next_elderly = int(pivot_elderly.loc[sido, base_ym_list[i+1]]) if pd.notna(pivot_elderly.loc[sido, base_ym_list[i+1]]) else 0
                    row[f'change_rate_{ym}'] = round((elderly - next_elderly) / next_elderly * 100, 2) if next_elderly > 0 else 0
        data.append(row)

    return {
        'headers': base_ym_list,
        'data': data,
        'description': elderly_def['description']
    }


def get_sigungu_population_table(base_ym_list, sido, include_sub_sigungu=False):
    """
    시군구별 인구 현황 테이블

    Args:
        base_ym_list: 기준년월 리스트
        sido: 시도명
        include_sub_sigungu: 하위 시군구 포함 여부 (예: 포항시 → 포항 북구/남구 분리)
    """
    engine = get_db_engine()

    if not sido:
        return {'headers': [], 'data': []}

    # 특별자치도 명칭 매핑 (과거/현재 명칭 모두 검색)
    sido_condition = f"sido_nm = '{sido}'"
    if sido == '강원특별자치도':
        sido_condition = "sido_nm IN ('강원특별자치도', '강원도')"
    elif sido == '전북특별자치도':
        sido_condition = "sido_nm IN ('전북특별자치도', '전라북도')"
    elif sido == '제주특별자치도':
        sido_condition = "sido_nm IN ('제주특별자치도', '제주도')"

    results = []
    for ym in base_ym_list:
        if include_sub_sigungu:
            # 하위 시군구 포함: 5자리 코드 전체 (구 레벨 분리)
            df = pd.read_sql(f"""
                SELECT
                    sigungu_code,
                    sigungu_nm,
                    SUM(total_pop) as total_pop,
                    SUM(COALESCE(household_cnt, 0)) as household_cnt,
                    SUM(COALESCE(elderly_pop, 0)) as elderly_pop,
                    ROUND(SUM(COALESCE(elderly_pop, 0))::numeric / NULLIF(SUM(total_pop), 0) * 100, 2) as elderly_ratio
                FROM cache_sigungu_indicators
                WHERE TO_CHAR(base_ym, 'YYYYMM') = '{ym}'
                  AND {sido_condition}
                  AND LENGTH(sigungu_code) = 5
                GROUP BY sigungu_code, sigungu_nm
                ORDER BY sigungu_code
            """, engine)
        else:
            # 하위 시군구 미포함: 앞 4자리로 그룹화 (포항 북구/남구 -> 포항시로 합산)
            df = pd.read_sql(f"""
                SELECT
                    LEFT(sigungu_code, 4) || '0' as sigungu_code,
                    COALESCE(
                        MAX(CASE WHEN sigungu_code LIKE '____0' THEN sigungu_nm END),
                        SPLIT_PART(MIN(sigungu_nm), ' ', 1)
                    ) as sigungu_nm,
                    SUM(total_pop) as total_pop,
                    SUM(COALESCE(household_cnt, 0)) as household_cnt,
                    SUM(COALESCE(elderly_pop, 0)) as elderly_pop,
                    ROUND(SUM(COALESCE(elderly_pop, 0))::numeric / NULLIF(SUM(total_pop), 0) * 100, 2) as elderly_ratio
                FROM cache_sigungu_indicators
                WHERE TO_CHAR(base_ym, 'YYYYMM') = '{ym}'
                  AND {sido_condition}
                  AND LENGTH(sigungu_code) = 5
                GROUP BY LEFT(sigungu_code, 4)
                ORDER BY LEFT(sigungu_code, 4)
            """, engine)
        df['base_ym'] = ym
        results.append(df)

    if not results:
        return {'headers': [], 'data': []}

    combined_df = pd.concat(results, ignore_index=True)

    if combined_df.empty:
        return {'headers': [], 'data': []}

    # sigungu_code를 인덱스로 사용하여 중복 방지
    pivot_pop = combined_df.pivot(index='sigungu_code', columns='base_ym', values='total_pop').fillna(0)
    pivot_household = combined_df.pivot(index='sigungu_code', columns='base_ym', values='household_cnt').fillna(0)
    pivot_elderly = combined_df.pivot(index='sigungu_code', columns='base_ym', values='elderly_ratio').fillna(0)

    # sigungu_code -> sigungu_nm 매핑
    code_to_name = combined_df.drop_duplicates('sigungu_code').set_index('sigungu_code')['sigungu_nm'].to_dict()

    data = []

    # 합계 행 추가
    total_row = {'sigungu_nm': '합계', 'sigungu_code': '00000'}
    for i, ym in enumerate(base_ym_list):
        if ym in pivot_pop.columns:
            total_row[f'pop_{ym}'] = int(pivot_pop[ym].sum())
            total_row[f'household_{ym}'] = int(pivot_household[ym].sum())
            # 고령화율은 가중평균
            total_elderly = combined_df[combined_df['base_ym'] == ym]['elderly_pop'].sum()
            total_pop_ym = combined_df[combined_df['base_ym'] == ym]['total_pop'].sum()
            total_row[f'elderly_rate_{ym}'] = round(total_elderly / total_pop_ym * 100, 1) if total_pop_ym > 0 else 0

            # 가장 오래된 시점(마지막)에는 증감률 표시 안함 - 직전 시점과 비교
            if i < len(base_ym_list) - 1 and base_ym_list[i+1] in pivot_pop.columns:
                next_pop = int(pivot_pop[base_ym_list[i+1]].sum())
                curr_pop = int(pivot_pop[ym].sum())
                total_row[f'pop_rate_{ym}'] = round((curr_pop - next_pop) / next_pop * 100, 2) if next_pop > 0 else 0
    data.append(total_row)

    # 각 시군구 행 (시군구코드 순)
    for code in pivot_pop.index:
        sigungu_nm = code_to_name.get(code, code)
        row = {'sigungu_nm': sigungu_nm, 'sigungu_code': code}
        for i, ym in enumerate(base_ym_list):
            if ym in pivot_pop.columns:
                row[f'pop_{ym}'] = int(pivot_pop.loc[code, ym])
                row[f'household_{ym}'] = int(pivot_household.loc[code, ym])
                row[f'elderly_rate_{ym}'] = round(pivot_elderly.loc[code, ym], 1)

                # 가장 오래된 시점(마지막)에는 증감률 표시 안함 - 직전 시점과 비교
                if i < len(base_ym_list) - 1 and base_ym_list[i+1] in pivot_pop.columns:
                    next_pop = int(pivot_pop.loc[code, base_ym_list[i+1]])
                    curr_pop = int(pivot_pop.loc[code, ym])
                    row[f'pop_rate_{ym}'] = round((curr_pop - next_pop) / next_pop * 100, 2) if next_pop > 0 else 0
        data.append(row)

    return {
        'headers': base_ym_list,
        'data': data
    }


def get_indicator_table(base_ym_list, indicator, aggregate_type='sido', region=None, sido=None, include_sub_sigungu=False):
    """
    동적 지표 데이터 조회 (code_indicator 기반)

    Args:
        base_ym_list: 기준년월 리스트
        indicator: 지표 정보 dict (column_name, display_name, description, numerator, ...)
        aggregate_type: 집계 단위 ('region', 'sido', 'sigungu')
        region: 권역명 (권역별일 때)
        sido: 시도명 (시군구별일 때)
        include_sub_sigungu: 하위 시군구(읍면동) 포함 여부

    Returns:
        dict: {'title': ..., 'description': ..., 'headers': [...], 'data': [...], 'numerator_name': ...}
    """
    engine = get_db_engine()
    col_name = indicator['column_name']
    display_name = indicator['display_name']
    description = indicator.get('description', '')
    numerator = indicator.get('numerator', '')  # 분자 컬럼명

    # 분자 컬럼이 없으면 기본값 설정
    if not numerator:
        numerator = 'total_pop'
        numerator_label = '인구'
    else:
        # DB에서 컬럼명 → 한글 라벨 매핑 조회
        numerator_labels = get_column_name_labels()
        # 복합 컬럼 처리 (예: youth_pop + elderly_pop -> 유소년인구 + 노령인구)
        if '+' in numerator:
            parts = [p.strip() for p in numerator.split('+')]
            translated_parts = [numerator_labels.get(p, p) for p in parts]
            numerator_label = ' + '.join(translated_parts)
        else:
            numerator_label = numerator_labels.get(numerator, numerator)

    # 집계 단위에 따른 쿼리 생성
    results = []
    group_label = '지역'  # 기본값

    for ym in base_ym_list:
        if aggregate_type == 'region':
            # 권역별: dim_admin_area에서 시군구별 region 정보만 DISTINCT로 조회하여 중복 조인 방지
            df = pd.read_sql(f"""
                SELECT
                    d.region_nm as name,
                    d.region_code,
                    SUM(COALESCE(c.{numerator}, 0)) as numerator_value,
                    SUM(c.total_pop) as total_pop,
                    ROUND(SUM(COALESCE(c.{numerator}, 0))::numeric / NULLIF(SUM(c.total_pop), 0) * 100, 2) as indicator_value
                FROM cache_sigungu_indicators c
                JOIN (
                    SELECT DISTINCT sigungu_code, region_nm, region_code
                    FROM dim_admin_area
                    WHERE region_nm IS NOT NULL
                ) d ON c.sigungu_code = d.sigungu_code
                WHERE TO_CHAR(c.base_ym, 'YYYYMM') = '{ym}'
                GROUP BY d.region_nm, d.region_code
                ORDER BY d.region_code
            """, engine)
            group_label = '권역'
        elif aggregate_type == 'sigungu' and sido:
            # 시군구별
            if sido == '강원특별자치도':
                sido_cond = "c.sido_nm IN ('강원특별자치도', '강원도')"
            elif sido == '전북특별자치도':
                sido_cond = "c.sido_nm IN ('전북특별자치도', '전라북도')"
            elif sido == '제주특별자치도':
                sido_cond = "c.sido_nm IN ('제주특별자치도', '제주도')"
            else:
                sido_cond = f"c.sido_nm = '{sido}'"

            # 하위 시군구(읍면동) 포함 여부에 따라 필터링
            sigungu_filter = "" if include_sub_sigungu else "AND c.sigungu_code LIKE '____0'"

            df = pd.read_sql(f"""
                SELECT
                    c.sigungu_code,
                    c.sigungu_nm as name,
                    SUM(COALESCE(c.{numerator}, 0)) as numerator_value,
                    SUM(c.total_pop) as total_pop,
                    ROUND(SUM(COALESCE(c.{numerator}, 0))::numeric / NULLIF(SUM(c.total_pop), 0) * 100, 2) as indicator_value
                FROM cache_sigungu_indicators c
                WHERE TO_CHAR(c.base_ym, 'YYYYMM') = '{ym}'
                  AND {sido_cond}
                  {sigungu_filter}
                GROUP BY c.sigungu_code, c.sigungu_nm
                ORDER BY c.sigungu_code
            """, engine)
            group_label = '시군' if not include_sub_sigungu else '시군구'
        else:
            # 시도별: dim_admin_area JOIN 없이 직접 조회 (과거 데이터 누락 방지)
            # region 필터가 있으면 먼저 해당 권역의 시도 목록을 조회
            region_cond = ""
            if region:
                # 권역에 해당하는 시도 목록 조회
                region_sidos = pd.read_sql(f"""
                    SELECT DISTINCT sido_nm
                    FROM dim_admin_area
                    WHERE region_nm = '{region}'
                """, engine)['sido_nm'].tolist()
                if region_sidos:
                    sidos_str = "', '".join(region_sidos)
                    region_cond = f"AND c.sido_nm IN ('{sidos_str}')"

            df = pd.read_sql(f"""
                SELECT
                    CASE
                        WHEN c.sido_nm IN ('강원특별자치도', '강원도') THEN '강원특별자치도'
                        WHEN c.sido_nm IN ('전북특별자치도', '전라북도') THEN '전북특별자치도'
                        WHEN c.sido_nm IN ('제주특별자치도', '제주도') THEN '제주특별자치도'
                        ELSE c.sido_nm
                    END as name,
                    MIN(LEFT(c.sigungu_code, 2)) as sido_code,
                    SUM(COALESCE(c.{numerator}, 0)) as numerator_value,
                    SUM(c.total_pop) as total_pop,
                    ROUND(SUM(COALESCE(c.{numerator}, 0))::numeric / NULLIF(SUM(c.total_pop), 0) * 100, 2) as indicator_value
                FROM cache_sigungu_indicators c
                WHERE TO_CHAR(c.base_ym, 'YYYYMM') = '{ym}'
                {region_cond}
                GROUP BY CASE
                    WHEN c.sido_nm IN ('강원특별자치도', '강원도') THEN '강원특별자치도'
                    WHEN c.sido_nm IN ('전북특별자치도', '전라북도') THEN '전북특별자치도'
                    WHEN c.sido_nm IN ('제주특별자치도', '제주도') THEN '제주특별자치도'
                    ELSE c.sido_nm END
                ORDER BY MIN(LEFT(c.sigungu_code, 2))
            """, engine)
            group_label = '시도'

        df['base_ym'] = ym
        results.append(df)

    if not results:
        return {'title': display_name, 'description': description, 'headers': [], 'data': [], 'group_label': group_label, 'numerator_name': numerator_label}

    combined_df = pd.concat(results, ignore_index=True)

    if combined_df.empty:
        return {'title': display_name, 'description': description, 'headers': [], 'data': [], 'group_label': group_label, 'numerator_name': numerator_label}

    # 코드순 정렬을 위해 코드 컬럼을 인덱스로 사용
    if aggregate_type == 'sigungu' and 'sigungu_code' in combined_df.columns:
        # 시군구코드를 인덱스로 사용
        pivot_value = combined_df.pivot(index='sigungu_code', columns='base_ym', values='indicator_value').fillna(0)
        pivot_numerator = combined_df.pivot(index='sigungu_code', columns='base_ym', values='numerator_value').fillna(0)
        pivot_denom = combined_df.pivot(index='sigungu_code', columns='base_ym', values='total_pop').fillna(0)
        # 시군구코드 순 정렬
        pivot_value = pivot_value.sort_index()
        pivot_numerator = pivot_numerator.sort_index()
        pivot_denom = pivot_denom.sort_index()
        # 시군구코드 -> 시군구명 매핑
        code_to_name = combined_df.drop_duplicates('sigungu_code').set_index('sigungu_code')['name'].to_dict()
    elif aggregate_type == 'region' and 'region_code' in combined_df.columns:
        # 권역코드를 인덱스로 사용
        pivot_value = combined_df.pivot(index='region_code', columns='base_ym', values='indicator_value').fillna(0)
        pivot_numerator = combined_df.pivot(index='region_code', columns='base_ym', values='numerator_value').fillna(0)
        pivot_denom = combined_df.pivot(index='region_code', columns='base_ym', values='total_pop').fillna(0)
        # 권역코드 순 정렬
        pivot_value = pivot_value.sort_index()
        pivot_numerator = pivot_numerator.sort_index()
        pivot_denom = pivot_denom.sort_index()
        # 권역코드 -> 권역명 매핑
        code_to_name = combined_df.drop_duplicates('region_code').set_index('region_code')['name'].to_dict()
    elif 'sido_code' in combined_df.columns:
        # 시도코드를 인덱스로 사용
        pivot_value = combined_df.pivot(index='sido_code', columns='base_ym', values='indicator_value').fillna(0)
        pivot_numerator = combined_df.pivot(index='sido_code', columns='base_ym', values='numerator_value').fillna(0)
        pivot_denom = combined_df.pivot(index='sido_code', columns='base_ym', values='total_pop').fillna(0)
        # 시도코드 순 정렬
        pivot_value = pivot_value.sort_index()
        pivot_numerator = pivot_numerator.sort_index()
        pivot_denom = pivot_denom.sort_index()
        # 시도코드 -> 시도명 매핑
        code_to_name = combined_df.drop_duplicates('sido_code').set_index('sido_code')['name'].to_dict()
    else:
        pivot_value = combined_df.pivot(index='name', columns='base_ym', values='indicator_value').fillna(0)
        pivot_numerator = combined_df.pivot(index='name', columns='base_ym', values='numerator_value').fillna(0)
        pivot_denom = combined_df.pivot(index='name', columns='base_ym', values='total_pop').fillna(0)
        code_to_name = None

    data = []

    # 전국/전체 합계
    total_row = {'name': '전국' if aggregate_type != 'sigungu' else '합계'}
    for i, ym in enumerate(base_ym_list):
        if ym in pivot_value.columns:
            # 분자 합계 및 인구 가중 평균 계산
            total_numerator = int(pivot_numerator[ym].sum())
            total_pop = int(pivot_denom[ym].sum())
            weighted_avg = round(total_numerator / total_pop * 100, 2) if total_pop > 0 else 0
            total_row[f'value_{ym}'] = weighted_avg
            total_row[f'numerator_{ym}'] = total_numerator
            total_row[f'denom_{ym}'] = total_pop
            # 가장 오래된 시점(마지막)에는 증감률 표시 안함 - 직전 시점과 비교
            if i < len(base_ym_list) - 1 and base_ym_list[i+1] in pivot_value.columns:
                next_ym = base_ym_list[i+1]  # 직전(더 오래된) 시점
                next_numerator = int(pivot_numerator[next_ym].sum())
                next_total_pop = int(pivot_denom[next_ym].sum())
                next_val = round(next_numerator / next_total_pop * 100, 2) if next_total_pop > 0 else 0
                total_row[f'change_{ym}'] = round(total_row[f'value_{ym}'] - next_val, 2)
    data.append(total_row)

    # 각 지역 행
    for idx in pivot_value.index:
        # 시군구별 조회 시 시군구코드 -> 시군구명 변환
        name = code_to_name.get(idx, idx) if code_to_name else idx
        row = {'name': name}
        for i, ym in enumerate(base_ym_list):
            if ym in pivot_value.columns:
                val = pivot_value.loc[idx, ym] if pd.notna(pivot_value.loc[idx, ym]) else 0
                numerator_val = int(pivot_numerator.loc[idx, ym]) if pd.notna(pivot_numerator.loc[idx, ym]) else 0
                denom_val = int(pivot_denom.loc[idx, ym]) if pd.notna(pivot_denom.loc[idx, ym]) else 0
                row[f'value_{ym}'] = round(val, 2)
                row[f'numerator_{ym}'] = numerator_val
                row[f'denom_{ym}'] = denom_val

                # 가장 오래된 시점(마지막)에는 증감률 표시 안함 - 직전 시점과 비교
                if i < len(base_ym_list) - 1 and base_ym_list[i+1] in pivot_value.columns:
                    next_val = pivot_value.loc[idx, base_ym_list[i+1]] if pd.notna(pivot_value.loc[idx, base_ym_list[i+1]]) else 0
                    row[f'change_{ym}'] = round(val - next_val, 2)
        data.append(row)

    return {
        'title': display_name,
        'description': description,
        'headers': base_ym_list,
        'data': data,
        'group_label': group_label,
        'numerator_name': numerator_label,
        'denominator_name': '총인구'
    }


def create_indicator_chart(data, ym_list, indicator_name):
    """
    지표별 차트 생성 (Matplotlib)
    - 지역별로 지표 값을 막대그래프로 표시
    - 최근 자료(첫 번째 년월)에만 레이블 표시
    """
    if not data or not data.get('data') or len(data['data']) < 2:
        return None

    try:
        # 전국 제외
        rows = data['data'][1:]
        labels = [r.get('name', '') for r in rows]
        num_items = len(labels)

        colors = ['#1D64F2', '#F24822', '#10b981', '#f59e0b', '#8b5cf6']

        # 항목이 10개 이상이면 가로 막대 그래프
        if num_items >= 10:
            fig_height = max(6, num_items * 0.35)
            fig, ax = plt.subplots(figsize=(10, fig_height))

            y = np.arange(len(labels))
            width = 0.8 / len(ym_list)

            for i, ym in enumerate(ym_list):
                key = f'value_{ym}'
                values = [r.get(key, 0) for r in rows]
                bars = ax.barh(y + i * width - (len(ym_list) - 1) * width / 2, values,
                       width, label=f'{ym[:4]}.{ym[4:]}', color=colors[i % len(colors)])
                # 최근 자료(첫 번째 년월)에만 레이블 표시
                if i == 0:
                    add_bar_labels(ax, bars, values, is_horizontal=True, format_str='{:.1f}')

            ax.set_ylabel('지역')
            ax.set_xlabel(indicator_name)
            ax.set_yticks(y)
            ax.set_yticklabels(labels)
            ax.legend(loc='lower right')
            ax.grid(axis='x', alpha=0.3)
            ax.invert_yaxis()
        else:
            fig, ax = plt.subplots(figsize=(12, 5))

            x = np.arange(len(labels))
            width = 0.8 / len(ym_list)

            for i, ym in enumerate(ym_list):
                key = f'value_{ym}'
                values = [r.get(key, 0) for r in rows]
                bars = ax.bar(x + i * width - (len(ym_list) - 1) * width / 2, values,
                       width, label=f'{ym[:4]}.{ym[4:]}', color=colors[i % len(colors)])
                # 최근 자료(첫 번째 년월)에만 레이블 표시
                if i == 0:
                    add_bar_labels(ax, bars, values, is_horizontal=False, format_str='{:.1f}')

            ax.set_xlabel('지역')
            ax.set_ylabel(indicator_name)
            ax.set_xticks(x)
            ax.set_xticklabels(labels, rotation=45, ha='right')
            ax.legend(loc='upper right')
            ax.grid(axis='y', alpha=0.3)

        plt.tight_layout()

        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=100, bbox_inches='tight')
        buf.seek(0)
        img_base64 = base64.b64encode(buf.read()).decode('utf-8')
        plt.close(fig)

        return img_base64
    except Exception as e:
        print(f"Indicator chart error: {e}")
        return None


def create_indicator_table_html(data, ym_list):
    """지표 데이터를 HTML 테이블로 변환 (지표값, 분자값, 증감)"""
    if not data or not data.get('data'):
        return '<p class="text-muted">데이터가 없습니다.</p>'

    rows = data['data']
    if not rows:
        return '<p class="text-muted">데이터가 없습니다.</p>'

    group_label = data.get('group_label', '지역')
    numerator_name = data.get('numerator_name', '인구')
    denominator_name = data.get('denominator_name', '총인구')

    html = ['<table class="data-table sortable-table">']

    # 헤더 행 (년월)
    html.append('<thead>')
    html.append('<tr>')
    html.append(f'<th rowspan="2" style="background:#1243A6;">{group_label}</th>')

    for i, ym in enumerate(ym_list):
        ym_display = f'{ym[:4]}년 {ym[4:]}월'
        divider_class = ' divider' if i > 0 else ''
        html.append(f'<th colspan="4" class="year-header{divider_class}">{ym_display}</th>')

    html.append('</tr>')

    # 두 번째 헤더 행 (지표값, 분자값, 분모값, 증감)
    html.append('<tr>')
    for i, ym in enumerate(ym_list):
        divider_class = ' divider' if i > 0 else ''
        html.append(f'<th class="metric-header sortable{divider_class}" data-col="value_{ym}">지표값 <span class="sort-icon"></span></th>')
        html.append(f'<th class="metric-header sortable" data-col="numerator_{ym}" style="font-size:0.85em;">{numerator_name} <span class="sort-icon"></span></th>')
        html.append(f'<th class="metric-header sortable" data-col="denom_{ym}" style="font-size:0.85em;">{denominator_name} <span class="sort-icon"></span></th>')
        html.append(f'<th class="metric-header sortable" data-col="change_{ym}">증감 <span class="sort-icon"></span></th>')

    html.append('</tr>')
    html.append('</thead>')

    # 데이터 행
    html.append('<tbody>')
    for row_idx, row in enumerate(rows):
        html.append('<tr>')
        html.append(f'<td>{row.get("name", "")}</td>')

        for i, ym in enumerate(ym_list):
            val = row.get(f'value_{ym}', 0)
            numerator_val = row.get(f'numerator_{ym}', 0)
            denom_val = row.get(f'denom_{ym}', 0)
            change = row.get(f'change_{ym}', None)

            divider_class = ' class="divider"' if i > 0 else ''
            html.append(f'<td{divider_class} data-value="{val}">{val:.2f}</td>')
            html.append(f'<td data-value="{numerator_val}" style="color:#666; font-size:0.9em;">{numerator_val:,}</td>')
            html.append(f'<td data-value="{denom_val}" style="color:#666; font-size:0.9em;">{denom_val:,}</td>')

            if change is not None:
                css_class = 'positive' if change > 0 else 'negative' if change < 0 else ''
                html.append(f'<td class="{css_class}" data-value="{change}">{change:+.2f}</td>')
            else:
                html.append('<td data-value="">-</td>')

        html.append('</tr>')

    html.append('</tbody>')
    html.append('</table>')

    return '\n'.join(html)


# =============================================================================
# 내보내기 함수들 (기존 코드 - 공통 모듈로 이전됨)
# =============================================================================
# [모듈화 완료]
# 아래 내보내기 함수들은 common/export_utils.py의 DataExporter 클래스로 이전되었습니다.
# - export_to_excel() -> DataExporter.export_to_excel()
# - export_to_markdown() -> DataExporter.export_to_markdown()
# - export_to_html() -> DataExporter.export_to_html()
#
# 사용 방법:
#   from common.export_utils import DataExporter
#   exporter = DataExporter(tables_data, charts_data, sort_config)
#   exporter.export_all(output_dir, filename)
#
# 이점:
# - 정렬 파라미터 지원 (화면 정렬 상태 반영)
# - 탭 형식 HTML 출력
# - 차트 이미지 포함
# - 다른 분야(경제, 환경 등) 데이터에도 재사용 가능
# =============================================================================

def _deprecated_export_to_excel(tables_data, output_path):
    """Excel 파일로 내보내기"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        wb = openpyxl.Workbook()

        header_fill = PatternFill(start_color='1243A6', end_color='1243A6', fill_type='solid')
        header_font_white = Font(bold=True, size=11, color='FFFFFF')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for idx, (sheet_name, table_data) in enumerate(tables_data.items()):
            if idx == 0:
                ws = wb.active
                ws.title = sheet_name[:31]
            else:
                ws = wb.create_sheet(title=sheet_name[:31])

            if not table_data.get('data'):
                continue

            # 제목 행
            start_row = 1
            if 'title' in table_data:
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
                ws.cell(row=1, column=1, value=table_data['title'])
                ws.cell(row=1, column=1).font = Font(bold=True, size=14)
                start_row = 3
            if 'category_name' in table_data and table_data['category_name']:
                ws.cell(row=start_row-1 if start_row > 1 else 1, column=1,
                        value=f"연령구분: {table_data['category_name']}")
                start_row = max(start_row, 3)

            # 헤더 행
            headers = list(table_data['data'][0].keys()) if table_data['data'] else []
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col, value=header)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            # 데이터 행
            for row_idx, row_data in enumerate(table_data['data'], start_row + 1):
                for col, header in enumerate(headers, 1):
                    value = row_data.get(header, '')
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = thin_border
                    if isinstance(value, (int, float)):
                        cell.alignment = Alignment(horizontal='right')

            # 열 너비 조정
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = min(max_length + 4, 35)

        wb.save(output_path)
        return True
    except Exception as e:
        print(f"Excel export error: {e}")
        return False


def _deprecated_export_to_markdown(tables_data, output_path):
    """Markdown 파일로 내보내기 (deprecated - common.export_utils 사용)"""
    try:
        lines = []
        lines.append("# 인구 집계표 보고서")
        lines.append(f"\n**생성일시:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        lines.append("---\n")

        for sheet_name, table_data in tables_data.items():
            lines.append(f"\n## {sheet_name}\n")

            if 'title' in table_data:
                lines.append(f"### {table_data['title']}\n")
            if 'category_name' in table_data and table_data['category_name']:
                lines.append(f"*연령구분: {table_data['category_name']}*\n")

            if not table_data.get('data'):
                lines.append("*데이터 없음*\n")
                continue

            headers = list(table_data['data'][0].keys())
            lines.append("| " + " | ".join(headers) + " |")
            lines.append("|" + "|".join(["---"] * len(headers)) + "|")

            for row in table_data['data']:
                values = []
                for h in headers:
                    v = row.get(h, '')
                    if isinstance(v, int):
                        values.append(f"{v:,}")
                    elif isinstance(v, float):
                        values.append(f"{v:.2f}")
                    else:
                        values.append(str(v))
                lines.append("| " + " | ".join(values) + " |")

            lines.append("")

        lines.append("\n---\n*본 보고서는 인구 집계표 대시보드에서 자동 생성되었습니다.*")

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))

        return True
    except Exception as e:
        print(f"Markdown export error: {e}")
        return False


def _deprecated_export_to_html(tables_data, output_path, charts_data=None):
    """
    [DEPRECATED] HTML 파일로 내보내기 (탭 형식, 차트 포함)
    이 함수는 common/export_utils.py의 DataExporter.export_to_html()로 대체됨
    정렬 파라미터 지원을 위해 DataExporter 클래스 사용을 권장
    """
    try:
        # 탭 ID 생성
        tab_items = []
        for idx, (sheet_name, table_data) in enumerate(tables_data.items()):
            tab_id = f"tab_{idx}"
            tab_items.append({
                'id': tab_id,
                'name': sheet_name.replace('_', ' '),
                'data': table_data,
                'chart': charts_data.get(sheet_name) if charts_data else None
            })

        html = """<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>인구 집계표 보고서</title>
    <style>
        * { box-sizing: border-box; }
        body { font-family: 'Pretendard', 'Noto Sans KR', sans-serif; margin: 0; padding: 20px; background: #f5f5f5; }
        .container { max-width: 1400px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
        h1 { color: #1243A6; border-bottom: 2px solid #1243A6; padding-bottom: 10px; margin-bottom: 20px; }
        .meta-info { color: #666; margin-bottom: 20px; }

        /* 탭 스타일 */
        .tab-nav { display: flex; flex-wrap: wrap; gap: 5px; border-bottom: 2px solid #1243A6; padding-bottom: 0; margin-bottom: 0; }
        .tab-btn {
            padding: 10px 20px; border: none; background: #e8f4fc; color: #1243A6;
            cursor: pointer; border-radius: 8px 8px 0 0; font-size: 14px; font-weight: 500;
            transition: all 0.2s;
        }
        .tab-btn:hover { background: #c5e3f6; }
        .tab-btn.active { background: #1243A6; color: white; }

        .tab-content { display: none; padding: 20px 0; }
        .tab-content.active { display: block; }

        /* 차트 */
        .chart-container { text-align: center; margin: 20px 0; }
        .chart-container img { max-width: 100%; height: auto; border: 1px solid #eee; border-radius: 8px; }

        /* 테이블 */
        .table-container { overflow-x: auto; }
        table { border-collapse: collapse; width: 100%; margin: 15px 0; font-size: 13px; }
        th { background: #1243A6; color: white; padding: 10px 8px; text-align: center; white-space: nowrap; }
        td { border: 1px solid #ddd; padding: 8px; text-align: right; }
        td:first-child { text-align: left; font-weight: 500; background: #f8f9fa; }
        tr:nth-child(even) { background: #f9f9f9; }
        tr:hover { background: #e8f4fc; }
        tr:first-child td { background: #e8f4fc; font-weight: bold; }

        .positive { color: #2563eb; }
        .negative { color: #dc2626; }

        .description { color: #666; font-style: italic; margin: 10px 0; padding: 10px; background: #f8f9fa; border-radius: 4px; }
        .footer { margin-top: 30px; padding-top: 20px; border-top: 1px solid #ddd; color: #666; font-size: 0.9em; }
    </style>
</head>
<body>
<div class="container">
"""
        html += f"<h1>인구 집계표 보고서</h1>"
        html += f"<p class='meta-info'><strong>생성일시:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>"

        # 탭 네비게이션
        html += '<div class="tab-nav">'
        for idx, tab in enumerate(tab_items):
            active_class = ' active' if idx == 0 else ''
            html += f'<button class="tab-btn{active_class}" onclick="showTab(\'{tab["id"]}\')">{tab["name"]}</button>'
        html += '</div>'

        # 탭 콘텐츠
        for idx, tab in enumerate(tab_items):
            active_class = ' active' if idx == 0 else ''
            html += f'<div id="{tab["id"]}" class="tab-content{active_class}">'

            table_data = tab['data']

            # 설명
            if 'description' in table_data and table_data['description']:
                html += f"<p class='description'>{table_data['description']}</p>"

            # 차트
            if tab['chart']:
                html += f'<div class="chart-container"><img src="data:image/png;base64,{tab["chart"]}" alt="{tab["name"]} 차트"></div>'

            # 테이블
            if table_data.get('data'):
                headers = list(table_data['data'][0].keys())
                # 코드 컬럼 제외
                headers = [h for h in headers if not h.endswith('_code') and h != 'sigungu_code']

                html += '<div class="table-container"><table><thead><tr>'
                for h in headers:
                    # 헤더명 한글화
                    h_display = h.replace('_', ' ')
                    if h.startswith('pop_'):
                        h_display = h.replace('pop_', '인구 ')
                    elif h.startswith('household_'):
                        h_display = h.replace('household_', '가구 ')
                    elif h.startswith('value_'):
                        h_display = h.replace('value_', '지표값 ')
                    elif h.startswith('numerator_'):
                        h_display = h.replace('numerator_', '분자 ')
                    elif h.startswith('change_'):
                        h_display = h.replace('change_', '증감 ')
                    elif h.startswith('rate_') or h.startswith('pop_rate_'):
                        h_display = h.replace('rate_', '증감률 ').replace('pop_rate_', '증감률 ')
                    elif h == 'name' or h == 'sigungu_nm':
                        h_display = '지역'
                    elif h == 'age_group':
                        h_display = '연령대'
                    html += f"<th>{h_display}</th>"
                html += "</tr></thead><tbody>"

                for row in table_data['data']:
                    html += "<tr>"
                    for h in headers:
                        val = row.get(h, '')
                        if isinstance(val, float):
                            css_class = ''
                            if 'rate' in h or 'change' in h:
                                css_class = 'positive' if val > 0 else 'negative' if val < 0 else ''
                            html += f"<td class='{css_class}'>{val:,.2f}</td>"
                        elif isinstance(val, int):
                            html += f"<td>{val:,}</td>"
                        else:
                            html += f"<td>{val}</td>"
                    html += "</tr>"

                html += "</tbody></table></div>"
            else:
                html += "<p><em>데이터 없음</em></p>"

            html += '</div>'

        # JavaScript
        html += """
<script>
function showTab(tabId) {
    // 모든 탭 콘텐츠 숨기기
    document.querySelectorAll('.tab-content').forEach(c => c.classList.remove('active'));
    // 모든 탭 버튼 비활성화
    document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
    // 선택된 탭 활성화
    document.getElementById(tabId).classList.add('active');
    // 해당 버튼 활성화
    event.target.classList.add('active');
}
</script>
"""

        html += """
<div class="footer">
    <p>본 보고서는 인구 집계표 대시보드에서 자동 생성되었습니다.</p>
</div>
</div>
</body>
</html>"""

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html)

        return True
    except Exception as e:
        print(f"HTML export error: {e}")
        import traceback
        traceback.print_exc()
        return False


# =============================================================================
# API 핸들러
# =============================================================================

def handle_api_request(api_type, request_args):
    """API 요청 처리"""

    if api_type == 'filter_options':
        return get_filter_options()

    # 다운로드 API (파일명만 필요, 다른 파라미터 불필요)
    if api_type == 'download':
        file_name = request_args.get('file', '')
        if not file_name:
            return {'error': '파일명이 지정되지 않았습니다.'}

        # 보안: 경로 조작 방지 (파일명만 허용)
        safe_name = Path(file_name).name
        file_path = POP_BASE / 'output' / safe_name

        if not file_path.exists():
            return {'error': f'파일을 찾을 수 없습니다: {safe_name}'}

        # 파일 확장자에 따른 MIME 타입 설정
        mime_types = {
            '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            '.md': 'text/markdown',
            '.html': 'text/html'
        }
        mime_type = mime_types.get(file_path.suffix.lower(), 'application/octet-stream')

        # send_file로 직접 반환
        return send_file(
            file_path,
            mimetype=mime_type,
            as_attachment=True,
            download_name=safe_name
        )

    # 공통 파라미터
    base_ym_str = request_args.get('base_ym_list', '')
    base_ym_list = sorted([ym.strip() for ym in base_ym_str.split(',') if ym.strip()], reverse=True)  # 최근순 정렬
    region = request_args.get('region')
    sido = request_args.get('sido')
    age_category = int(request_args.get('age_category', 2))  # 기본값: 10세별
    aggregate_type = request_args.get('aggregate_type', 'sido')
    include_sub_sigungu = request_args.get('include_sub_sigungu', '0') == '1'

    if not base_ym_list:
        return {'error': '기준년월을 선택해주세요.'}

    if api_type == 'age_table':
        return get_age_population_table(base_ym_list, age_category)

    elif api_type == 'sido_table':
        return get_sido_population_table(base_ym_list, region)

    elif api_type == 'sigungu_table':
        return get_sigungu_population_table(base_ym_list, sido, include_sub_sigungu)

    elif api_type == 'export':
        # ============================================================================
        # [내보내기 API] 정렬 파라미터를 포함한 데이터 내보내기
        # - sort_column: 정렬할 컬럼명 (예: 'pop_202512', 'value_202512')
        # - sort_direction: 정렬 방향 ('asc' 또는 'desc')
        # - 화면에서 정렬한 상태가 내보내기에 반영됨
        # ============================================================================
        export_format = request_args.get('format', 'all')

        # [정렬 파라미터] 화면에서 정렬한 컬럼과 방향을 받아옴
        sort_column = request_args.get('sort_column', '')
        sort_direction = request_args.get('sort_direction', 'asc')
        sort_config = {'column': sort_column, 'direction': sort_direction} if sort_column else None

        # 기본 테이블 데이터 조회
        age_data = get_age_population_table(base_ym_list, age_category)
        region_data = get_sido_population_table(base_ym_list, region)

        tables_data = {
            '연령별_인구현황': age_data,
            '지역별_인구현황': region_data,
        }

        # 차트 데이터 생성 (base64 인코딩된 이미지)
        charts_data = {
            '연령별_인구현황': create_age_chart(age_data, base_ym_list),
            '지역별_인구현황': create_region_chart(region_data, base_ym_list),
        }

        # 시군구별 데이터 (시도 선택 시)
        if sido:
            sigungu_data = get_sigungu_population_table(base_ym_list, sido, include_sub_sigungu)
            tables_data['시군구별_인구현황'] = sigungu_data
            charts_data['시군구별_인구현황'] = create_region_chart(sigungu_data, base_ym_list)

        # 동적 지표 테이블 (code_indicator 기반)
        active_indicators = get_active_indicators(1)  # 인구지표
        for indicator in active_indicators:
            display_name = indicator['display_name'].replace(' ', '_')
            indicator_data = get_indicator_table(
                base_ym_list, indicator, aggregate_type, region, sido, include_sub_sigungu
            )
            tables_data[display_name] = indicator_data
            charts_data[display_name] = create_indicator_chart(indicator_data, base_ym_list, indicator['display_name'])

        # ============================================================================
        # [모듈화] DataExporter 클래스 사용 - 다른 분야에서도 재사용 가능
        # - 정렬, 차트, 탭 형식 등을 지원하는 범용 내보내기 클래스
        # ============================================================================
        output_dir = POP_BASE / 'output'
        output_dir.mkdir(exist_ok=True)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        region_str = region or '전국'
        sido_str = sido or ''
        filename_base = f"인구집계표_{region_str}_{sido_str}_{timestamp}".replace(' ', '_')

        # DataExporter를 사용한 내보내기 (정렬 파라미터 전달)
        exporter = DataExporter(
            tables_data=tables_data,
            charts_data=charts_data,
            sort_config=sort_config,  # 화면 정렬 상태 반영
            report_title="인구 집계표 보고서"
        )

        results = exporter.export_all(output_dir, filename_base)

        # 다운로드 URL 생성 (클라우드 서버 배포용)
        if results.get('success') and results.get('files'):
            download_links = []
            for file_path in results['files']:
                file_name = Path(file_path).name
                # 현재 요청 URL의 기본 경로를 사용하여 다운로드 URL 생성
                download_url = f"?api_type=download&file={file_name}"
                file_ext = Path(file_path).suffix.upper().replace('.', '')
                download_links.append({
                    'name': file_name,
                    'url': download_url,
                    'type': file_ext
                })
            results['download_links'] = download_links

        return results

    return {'error': f'Unknown api_type: {api_type}'}


# =============================================================================
# Matplotlib 차트 생성 함수들
# =============================================================================

def add_bar_labels(ax, bars, values, is_horizontal=False, fontsize=8, format_str='{:.1f}', offset=0.3):
    """
    막대 그래프에 레이블 값을 추가하는 공통 함수 (모듈화)

    Args:
        ax: matplotlib axes 객체
        bars: bar 객체 리스트
        values: 값 리스트
        is_horizontal: 가로 막대 그래프 여부
        fontsize: 폰트 크기
        format_str: 값 포맷 문자열 ('{:.1f}', '{:,.0f}' 등)
        offset: 레이블과 막대 사이 간격
    """
    for bar, val in zip(bars, values):
        if val > 0:
            if is_horizontal:
                # 가로 막대 그래프
                ax.text(bar.get_width() + offset, bar.get_y() + bar.get_height()/2,
                       format_str.format(val), va='center', ha='left', fontsize=fontsize)
            else:
                # 세로 막대 그래프
                ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + offset,
                       format_str.format(val), ha='center', va='bottom', fontsize=fontsize)


def create_age_chart(data, ym_list):
    """
    연령별 인구 차트 생성 (Matplotlib)
    - 항목이 12개 이상이면 가로 막대 그래프로 변환
    - 최근 자료(첫 번째 년월)에만 레이블 표시
    """
    if not data or not data.get('data') or len(data['data']) < 2:
        return None

    try:
        # 총인구 제외, 모든 연령대
        rows = data['data'][1:]
        labels = [r.get('age_group', '') for r in rows]
        num_items = len(labels)

        colors = ['#1D64F2', '#F24822', '#10b981', '#f59e0b', '#8b5cf6']

        # 항목이 12개 이상이면 가로 막대 그래프
        if num_items >= 12:
            fig_height = max(8, num_items * 0.4)
            fig, ax = plt.subplots(figsize=(12, fig_height))

            y = np.arange(len(labels))
            width = 0.8 / len(ym_list)

            for i, ym in enumerate(ym_list):
                key = f'pop_{ym}'
                values = [r.get(key, 0) / 10000 for r in rows]  # 만 단위
                bars = ax.barh(y + i * width - (len(ym_list) - 1) * width / 2, values,
                       width, label=f'{ym[:4]}.{ym[4:]}', color=colors[i % len(colors)])
                # 최근 자료(첫 번째 년월)에만 레이블 표시
                if i == 0:
                    add_bar_labels(ax, bars, values, is_horizontal=True, format_str='{:,.0f}')

            ax.set_ylabel('연령대')
            ax.set_xlabel('인구 (만 명)')
            ax.set_yticks(y)
            ax.set_yticklabels(labels)
            ax.legend(loc='lower right')
            ax.grid(axis='x', alpha=0.3)
            ax.invert_yaxis()  # 0-4세가 위로 오도록
        else:
            # 항목이 적으면 세로 막대 그래프
            fig, ax = plt.subplots(figsize=(12, 5))

            x = np.arange(len(labels))
            width = 0.8 / len(ym_list)

            for i, ym in enumerate(ym_list):
                key = f'pop_{ym}'
                values = [r.get(key, 0) / 10000 for r in rows]  # 만 단위
                bars = ax.bar(x + i * width - (len(ym_list) - 1) * width / 2, values,
                       width, label=f'{ym[:4]}.{ym[4:]}', color=colors[i % len(colors)])
                # 최근 자료(첫 번째 년월)에만 레이블 표시
                if i == 0:
                    add_bar_labels(ax, bars, values, is_horizontal=False, format_str='{:,.0f}')

            ax.set_xlabel('연령대')
            ax.set_ylabel('인구 (만 명)')
            ax.set_xticks(x)
            ax.set_xticklabels(labels, rotation=45, ha='right')
            ax.legend(loc='upper right')
            ax.grid(axis='y', alpha=0.3)

        plt.tight_layout()

        # Base64로 변환
        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=100, bbox_inches='tight')
        buf.seek(0)
        img_base64 = base64.b64encode(buf.read()).decode('utf-8')
        plt.close(fig)

        return img_base64
    except Exception as e:
        print(f"Age chart error: {e}")
        return None


def create_age_chart_by_region(regional_data, ym_list):
    """
    지역별 연령 인구 subplot 차트 생성

    Args:
        regional_data: {지역명: {'total_pop': {ym: pop}, 'age_data': [{age_group: ..., pop_YYYYMM: ...}, ...]}, ...}
        ym_list: 기준년월 리스트
    """
    if not regional_data:
        return None

    try:
        # 설정에서 가져오기
        max_regions = REGIONAL_SUBPLOT.get('max_regions')  # None이면 전체
        n_cols_config = REGIONAL_SUBPLOT.get('cols', 3)
        fig_w = REGIONAL_SUBPLOT.get('fig_width_per_col', 5)
        fig_h = REGIONAL_SUBPLOT.get('fig_height_per_row', 4)
        unit_val = POPULATION_UNIT['unit']
        unit_label = POPULATION_UNIT['label']

        # 전체 지역 또는 max_regions 개수만큼
        regions = list(regional_data.keys())
        if max_regions is not None:
            regions = regions[:max_regions]
        n_regions = len(regions)

        if n_regions == 0:
            return None

        # subplot 레이아웃 결정
        n_cols = min(n_cols_config, n_regions)
        n_rows = (n_regions + n_cols - 1) // n_cols

        fig, axes = plt.subplots(n_rows, n_cols, figsize=(fig_w * n_cols, fig_h * n_rows))

        # axes를 2D 배열로 변환
        if n_regions == 1:
            axes = np.array([[axes]])
        elif n_rows == 1:
            axes = axes.reshape(1, -1)
        elif n_cols == 1:
            axes = axes.reshape(-1, 1)

        colors = ['#1D64F2', '#F24822', '#10b981', '#f59e0b', '#8b5cf6']

        for idx, region_name in enumerate(regions):
            row_idx = idx // n_cols
            col_idx = idx % n_cols
            ax = axes[row_idx, col_idx]

            # 새 데이터 구조: regional_data[region_name]['age_data']
            region_info = regional_data[region_name]
            rows = region_info.get('age_data', region_info) if isinstance(region_info, dict) else region_info
            labels = [r.get('age_group', '') for r in rows]

            x = np.arange(len(labels))
            width = 0.8 / len(ym_list)

            for i, ym in enumerate(ym_list):
                key = f'pop_{ym}'
                values = [r.get(key, 0) / unit_val for r in rows]
                bars = ax.bar(x + i * width - (len(ym_list) - 1) * width / 2, values,
                       width, label=f'{ym[:4]}.{ym[4:]}', color=colors[i % len(colors)])

                # 첫 번째 년월(가장 최근)에만 레이블 표시
                if i == 0:
                    for bar, val in zip(bars, values):
                        if val > 0:
                            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height(),
                                   f'{val:.0f}', ha='center', va='bottom',
                                   fontsize=8, fontweight='bold', color='#333')

            ax.set_title(region_name, fontsize=13, fontweight='bold')
            ax.set_xlabel('연령대', fontsize=10, fontweight='bold')
            ax.set_ylabel(f'인구 ({unit_label})', fontsize=10, fontweight='bold')
            ax.set_xticks(x)
            ax.set_xticklabels(labels, rotation=45, ha='right', fontsize=9, fontweight='bold')
            ax.tick_params(axis='y', labelsize=9)
            ax.grid(axis='y', alpha=0.3)

            # 첫 번째 subplot에만 범례 표시
            if idx == 0:
                ax.legend(loc='upper right', fontsize=9, fontweight='normal')

        # 빈 subplot 숨기기
        total_subplots = n_rows * n_cols
        for idx in range(n_regions, total_subplots):
            row_idx = idx // n_cols
            col_idx = idx % n_cols
            axes[row_idx, col_idx].set_visible(False)

        plt.tight_layout()

        # Base64로 변환
        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=100, bbox_inches='tight')
        buf.seek(0)
        img_base64 = base64.b64encode(buf.read()).decode('utf-8')
        plt.close(fig)

        return img_base64
    except Exception as e:
        print(f"Age chart by region error: {e}")
        import traceback
        traceback.print_exc()
        return None


def create_region_chart(data, ym_list, chart_type='sido'):
    """지역별 인구 차트 - 연도별 막대 그래프 (Matplotlib)"""
    if not data or not data.get('data') or len(data['data']) < 2:
        return None

    try:
        # 전국/합계 제외하고, 이름이 있는 항목만 필터링
        raw_rows = data['data'][1:]
        rows = []
        labels = []
        for r in raw_rows:
            name = r.get('name', r.get('sigungu_nm', r.get('sido_nm', '')))
            if name and name.strip():  # 빈 이름 제외
                rows.append(r)
                labels.append(name)
        num_items = len(labels)

        colors = ['#1D64F2', '#F24822', '#10b981', '#f59e0b', '#8b5cf6']

        # 항목이 10개 이상이면 가로 막대 그래프
        if num_items >= 10:
            fig_height = max(6, num_items * 0.35)
            fig, ax = plt.subplots(figsize=(10, fig_height))

            y = np.arange(len(labels))
            width = 0.8 / len(ym_list)

            for i, ym in enumerate(ym_list):
                pop_key = f'pop_{ym}'
                values = [r.get(pop_key, 0) / 10000 for r in rows]  # 만 단위
                bars = ax.barh(y + i * width - (len(ym_list) - 1) * width / 2, values,
                       width, label=f'{ym[:4]} {ym[4:]}', color=colors[i % len(colors)])
                # 최근 자료(첫 번째 년월)에만 레이블 표시
                if i == 0:
                    add_bar_labels(ax, bars, values, is_horizontal=True, format_str='{:,.0f}', offset=0.5)

            ax.set_ylabel('지역')
            ax.set_xlabel('인구 (만 명)')
            ax.set_yticks(y)
            ax.set_yticklabels(labels)
            ax.legend(loc='lower right')
            ax.grid(axis='x', alpha=0.3)
            ax.invert_yaxis()
        else:
            fig, ax = plt.subplots(figsize=(12, 5))

            x = np.arange(len(labels))
            width = 0.8 / len(ym_list)

            for i, ym in enumerate(ym_list):
                pop_key = f'pop_{ym}'
                values = [r.get(pop_key, 0) / 10000 for r in rows]  # 만 단위
                bars = ax.bar(x + i * width - (len(ym_list) - 1) * width / 2, values,
                       width, label=f'{ym[:4]} {ym[4:]}', color=colors[i % len(colors)])
                # 최근 자료(첫 번째 년월)에만 레이블 표시
                if i == 0:
                    add_bar_labels(ax, bars, values, is_horizontal=False, format_str='{:,.0f}', offset=0.5)

            ax.set_xlabel('지역')
            ax.set_ylabel('인구 (만 명)')
            ax.set_xticks(x)
            ax.set_xticklabels(labels, rotation=45, ha='right')
            ax.legend(loc='upper right')
            ax.grid(axis='y', alpha=0.3)

        plt.tight_layout()

        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=100, bbox_inches='tight')
        buf.seek(0)
        img_base64 = base64.b64encode(buf.read()).decode('utf-8')
        plt.close(fig)

        return img_base64
    except Exception as e:
        print(f"Region chart error: {e}")
        return None


def create_elderly_chart(data, ym_list):
    """노령인구 차트 (Matplotlib)"""
    return create_region_chart(data, ym_list, 'elderly')


def create_youth_chart(data, ym_list):
    """청년인구 차트 (Matplotlib)"""
    if not data or not data.get('data') or len(data['data']) < 2:
        return None

    try:
        fig, ax = plt.subplots(figsize=(12, 5))

        rows = data['data'][1:]
        labels = [r.get('sido_nm', '') for r in rows]
        x = np.arange(len(labels))
        width = 0.8 / len(ym_list)

        colors = ['#1D64F2', '#F24822', '#10b981', '#f59e0b']

        for i, ym in enumerate(ym_list):
            key = f'pop_{ym}'
            values = [r.get(key, 0) / 10000 for r in rows]
            ax.bar(x + i * width - (len(ym_list) - 1) * width / 2, values,
                   width, label=f'{ym[:4]}.{ym[4:]}', color=colors[i % len(colors)])

        ax.set_xlabel('시도')
        ax.set_ylabel('청년인구 (만 명)')
        ax.set_xticks(x)
        ax.set_xticklabels(labels, rotation=45, ha='right')
        ax.legend(loc='upper right')
        ax.grid(axis='y', alpha=0.3)
        plt.tight_layout()

        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=100, bbox_inches='tight')
        buf.seek(0)
        img_base64 = base64.b64encode(buf.read()).decode('utf-8')
        plt.close(fig)

        return img_base64
    except Exception as e:
        print(f"Youth chart error: {e}")
        return None


# =============================================================================
# 테이블 HTML 생성 (2줄 헤더 지원)
# =============================================================================

def create_age_by_region_table_html(regional_data, ym_list, max_regions=None, start_index=0):
    """
    지역별 연령 인구 테이블 HTML 생성 (접이식)

    Args:
        regional_data: {지역명: {'total_pop': {ym: pop}, 'age_data': [...]}, ...}
        ym_list: 기준년월 리스트
        max_regions: 표시할 최대 지역 수 (None이면 전체)
        start_index: 시작 인덱스 (더보기 기능용)
    """
    if not regional_data:
        return '<p class="text-muted">지역별 연령 데이터가 없습니다.</p>'

    # 설정에서 가져오기
    expand_all = ACCORDION_TABLE.get('expand_all', True)

    html = []
    # expand_all이면 data-bs-parent 제거 (각 아코디언 독립적으로 열림)
    parent_attr = '' if expand_all else 'data-bs-parent="#regionAgeAccordion"'
    html.append('<div class="accordion" id="regionAgeAccordion">')

    regions = list(regional_data.keys())
    total_regions = len(regions)

    # max_regions가 지정되면 해당 개수만큼만 표시
    if max_regions:
        end_index = min(start_index + max_regions, total_regions)
        regions = regions[start_index:end_index]

    for idx, region_name in enumerate(regions):
        region_info = regional_data[region_name]

        # 새 데이터 구조 지원
        if isinstance(region_info, dict) and 'age_data' in region_info:
            age_rows = region_info['age_data']
            total_pop = region_info.get('total_pop', {})
        else:
            age_rows = region_info
            total_pop = {}

        collapse_id = f"collapseRegion{start_index + idx}"
        heading_id = f"headingRegion{start_index + idx}"

        # 설정에 따라 전체 펼침 또는 첫 번째만 펼침
        if expand_all:
            collapsed = ''
            show = 'show'
            expanded = 'true'
        else:
            collapsed = '' if idx == 0 else 'collapsed'
            show = 'show' if idx == 0 else ''
            expanded = 'true' if idx == 0 else 'false'

        # 아코디언 아이템
        html.append(f'''
        <div class="accordion-item">
            <h2 class="accordion-header" id="{heading_id}">
                <button class="accordion-button {collapsed}" type="button"
                        data-bs-toggle="collapse" data-bs-target="#{collapse_id}"
                        aria-expanded="{expanded}" aria-controls="{collapse_id}">
                    <strong>{region_name}</strong>
                </button>
            </h2>
            <div id="{collapse_id}" class="accordion-collapse collapse {show}"
                 aria-labelledby="{heading_id}" {parent_attr}>
                <div class="accordion-body p-2">
        ''')

        # 테이블
        html.append('<table class="data-table">')

        # 헤더 - 2줄 구조
        html.append('<thead>')
        # 1줄: 년월 헤더
        html.append('<tr>')
        html.append(f'<th rowspan="2" style="min-width: 100px;">구분</th>')
        for i, ym in enumerate(ym_list):
            year = ym[:4]
            month = ym[4:]
            divider_class = ' divider' if i > 0 else ''
            html.append(f'<th class="year-header{divider_class}">{year}년 {month}월</th>')
        html.append('</tr>')

        # 2줄: 인구 헤더
        html.append('<tr>')
        for i, ym in enumerate(ym_list):
            divider_class = ' divider' if i > 0 else ''
            html.append(f'<th class="metric-header{divider_class}">인구</th>')
        html.append('</tr>')
        html.append('</thead>')

        # 바디
        html.append('<tbody>')

        # 총인구 행 추가 (맨 위)
        if total_pop:
            html.append('<tr style="background-color: #f0f4ff; font-weight: bold;">')
            html.append('<td>총인구</td>')
            for i, ym in enumerate(ym_list):
                pop = total_pop.get(ym, 0)
                divider_class = ' class="divider"' if i > 0 else ''
                html.append(f'<td{divider_class}>{pop:,}</td>')
            html.append('</tr>')

        # 연령별 행
        for row in age_rows:
            age_group = row.get('age_group', '')
            html.append('<tr>')
            html.append(f'<td>{age_group}</td>')
            for i, ym in enumerate(ym_list):
                pop = row.get(f'pop_{ym}', 0)
                divider_class = ' class="divider"' if i > 0 else ''
                html.append(f'<td{divider_class}>{pop:,}</td>')
            html.append('</tr>')
        html.append('</tbody>')
        html.append('</table>')

        html.append('</div></div></div>')

    html.append('</div>')

    # 더보기 버튼 (남은 지역이 있는 경우)
    if max_regions and (start_index + max_regions) < total_regions:
        remaining = total_regions - (start_index + max_regions)
        html.append(f'''
        <div class="text-center mt-3">
            <button type="button" class="btn btn-outline-primary btn-sm" id="loadMoreRegionsBtn"
                    onclick="loadMoreRegions({start_index + max_regions}, {max_regions})">
                더보기 ({remaining}개 지역 더)
            </button>
        </div>
        ''')

    return '\n'.join(html)


def create_table_html(data, ym_list, table_type='age'):
    """DataFrame을 2줄 헤더 HTML 테이블로 변환 (정렬 기능 포함)"""
    if not data or not data.get('data'):
        return '<p class="text-muted">데이터가 없습니다.</p>'

    rows = data['data']
    if not rows:
        return '<p class="text-muted">데이터가 없습니다.</p>'

    # 첫 번째 컬럼명 (행 이름)
    if table_type == 'age':
        first_col = 'age_group'
        first_col_name = '연령대'
    elif 'sigungu_nm' in rows[0]:
        first_col = 'sigungu_nm'
        first_col_name = '시군구'
    elif 'name' in rows[0]:
        first_col = 'name'
        first_col_name = '지역'
    else:
        first_col = 'sido_nm'
        first_col_name = '시도'

    # 지표 정의
    if table_type == 'age':
        metrics = [('pop', '인구'), ('rate', '증감률')]
    elif table_type == 'elderly':
        metrics = [('pop', '인구'), ('aging_rate', '고령화율'), ('change_rate', '증감률')]
    elif table_type == 'youth':
        metrics = [('pop', '인구'), ('rate', '증감률')]
    else:  # region (sido)
        metrics = [('pop', '인구'), ('household', '가구'), ('pop_rate', '증감률')]

    # HTML 생성
    html = ['<table class="data-table sortable-table">']

    # 첫 번째 헤더 행 (년월)
    html.append('<thead>')
    html.append('<tr>')
    html.append(f'<th rowspan="2" style="background:#1243A6;">{first_col_name}</th>')

    for i, ym in enumerate(ym_list):
        ym_display = f'{ym[:4]}년 {ym[4:]}월'
        divider_class = ' class="divider"' if i > 0 else ''
        colspan = len([m for m in metrics if any(f'{m[0]}_{ym}' in r for r in rows)])
        if colspan == 0:
            colspan = len(metrics)
        html.append(f'<th colspan="{colspan}"{divider_class} class="year-header">{ym_display}</th>')

    html.append('</tr>')

    # 두 번째 헤더 행 (지표) - 정렬 기능 추가
    html.append('<tr>')
    for i, ym in enumerate(ym_list):
        for j, (metric_key, metric_name) in enumerate(metrics):
            key = f'{metric_key}_{ym}'
            if any(key in r for r in rows):
                divider_class = ' divider' if i > 0 and j == 0 else ''
                html.append(f'<th class="metric-header sortable{divider_class}" data-col="{key}">{metric_name} <span class="sort-icon"></span></th>')

    html.append('</tr>')
    html.append('</thead>')

    # 데이터 행
    html.append('<tbody>')
    for row_idx, row in enumerate(rows):
        html.append('<tr>')
        html.append(f'<td>{row.get(first_col, "")}</td>')

        for i, ym in enumerate(ym_list):
            for j, (metric_key, metric_name) in enumerate(metrics):
                key = f'{metric_key}_{ym}'
                if key in row:
                    val = row[key]
                    divider_class = ' class="divider"' if i > 0 and j == 0 else ''

                    if isinstance(val, (int, float)):
                        if 'rate' in key:
                            css_class = 'positive' if val > 0 else 'negative' if val < 0 else ''
                            val_str = f'{val:.2f}%'
                            html.append(f'<td{divider_class} class="{css_class}" data-value="{val}">{val_str}</td>')
                        else:
                            val_str = f'{val:,}'
                            html.append(f'<td{divider_class} data-value="{val}">{val_str}</td>')
                    else:
                        html.append(f'<td{divider_class} data-value="{val}">{val}</td>')

    html.append('</tr>')
    html.append('</tbody>')
    html.append('</table>')

    return '\n'.join(html)


# =============================================================================
# Jinja2 템플릿 기반 HTML 생성
# =============================================================================

def generate_dashboard_html(request_args):
    """Jinja2 템플릿을 사용하여 대시보드 HTML 생성"""
    # 파라미터 파싱
    base_ym_str = request_args.get('base_ym_list', '')
    base_ym_list = sorted([ym.strip() for ym in base_ym_str.split(',') if ym.strip()], reverse=True)  # 최근순 정렬
    age_category = int(request_args.get('age_category', 2))  # 기본값: 10세별
    aggregate_type = request_args.get('aggregate_type', 'sido')
    sido = request_args.get('sido', '')
    region = request_args.get('region', '')
    active_tab = request_args.get('active_tab', 'region')
    include_sub_sigungu = request_args.get('include_sub_sigungu', '0') == '1'
    show_all_regions = request_args.get('show_all_regions', '0') == '1'

    # 필터 옵션 로드
    filters = get_filter_options()

    # 초기 접속 시 (base_ym_list가 없으면) 데이터 로드하지 않음 - 로딩 속도 최적화
    if not base_ym_list:
        # 기본 년월 설정 (UI 표시용, 실제 데이터 조회는 안함)
        dec_list = sorted([ym for ym in filters['base_ym_list'] if ym.endswith('12')], reverse=True)
        default_ym_list = dec_list[:4] if len(dec_list) >= 4 else sorted(filters['base_ym_list'], reverse=True)[:3]

        # 메뉴 생성
        menu_items = MenuGenerator.get_category_menu_items(POP_BASE, '01_population')

        # 초기 화면: 필터만 표시, 데이터 없음
        template = jinja_env.get_template('edu_dash2.html')
        return template.render(
            filters=filters,
            selected_ym_list=default_ym_list,
            selected_age_category=age_category,
            aggregate_type=aggregate_type,
            selected_sido=sido,
            selected_region=region,
            active_tab=active_tab,
            include_sub_sigungu=include_sub_sigungu,
            menu_items=menu_items,
            current_url='/01_population/edu_dash2',
            # 데이터 없음 (초기 접속)
            age_chart_img=None,
            region_chart_img=None,
            age_by_region_chart_img=None,
            age_table_html='<p class="text-muted text-center py-5">조회 버튼을 클릭하여 데이터를 불러오세요.</p>',
            region_table_html='<p class="text-muted text-center py-5">조회 버튼을 클릭하여 데이터를 불러오세요.</p>',
            age_by_region_table_html='',
            indicator_tabs=[{
                'id': ind['column_name'],
                'name': ind['display_name'],
                'description': ind.get('description', ''),
                'chart_img': None,
                'table_html': '<p class="text-muted text-center py-5">조회 버튼을 클릭하여 데이터를 불러오세요.</p>',
                'data': None
            } for ind in filters.get('active_indicators', [])],
            age_data=None,
            age_by_region_data=None,
            is_initial_load=True
        )

    # 데이터 조회 (조회 버튼 클릭 시)
    age_data = get_age_population_table(base_ym_list, age_category)

    # 집계 단위에 따른 지역 데이터 조회
    if aggregate_type == 'sigungu' and sido:
        # 시군구별: 해당 시도의 시군구 데이터
        region_data = get_sigungu_population_table(base_ym_list, sido, include_sub_sigungu)
    elif aggregate_type == 'region':
        # 권역별: 수도권, 경상권 등으로 그룹화
        region_data = get_region_population_table(base_ym_list)
    else:
        # 시도별: 전국 시도 데이터
        region_data = get_sido_population_table(base_ym_list, None)

    # 동적 지표 탭 데이터 조회 (code_indicator 기반)
    active_indicators = filters.get('active_indicators', [])
    indicator_tabs = []

    for indicator in active_indicators:
        col_name = indicator['column_name']
        display_name = indicator['display_name']
        description = indicator.get('description', '')

        # 지표 데이터 조회
        indicator_data = get_indicator_table(
            base_ym_list, indicator, aggregate_type, region, sido, include_sub_sigungu
        )

        # 차트 생성
        indicator_chart = create_indicator_chart(indicator_data, base_ym_list, display_name)

        # 테이블 HTML 생성
        indicator_table = create_indicator_table_html(indicator_data, base_ym_list)

        indicator_tabs.append({
            'id': col_name,
            'name': display_name,
            'description': description,
            'chart_img': indicator_chart,
            'table_html': indicator_table,
            'data': indicator_data
        })

    # 기본 차트 생성
    age_chart_img = create_age_chart(age_data, base_ym_list)
    region_chart_img = create_region_chart(region_data, base_ym_list)

    # 지역별 연령 인구 차트 (subplot)
    age_by_region_data = get_age_population_by_region(base_ym_list, age_category, aggregate_type, None)
    age_by_region_chart_img = create_age_chart_by_region(age_by_region_data, base_ym_list)
    # max_regions: 6개만 표시, show_all_regions가 true면 전체 표시
    max_regions_display = None if show_all_regions else 6
    age_by_region_table_html = create_age_by_region_table_html(age_by_region_data, base_ym_list, max_regions=max_regions_display)

    # 기본 테이블 HTML 생성
    age_table_html = create_table_html(age_data, base_ym_list, 'age')
    region_table_html = create_table_html(region_data, base_ym_list, 'region')

    # 메뉴 생성
    menu_items = MenuGenerator.get_category_menu_items(POP_BASE, '01_population')

    # 템플릿 렌더링
    template = jinja_env.get_template('edu_dash2.html')
    return template.render(
        filters=filters,
        selected_ym_list=base_ym_list,
        selected_age_category=age_category,
        aggregate_type=aggregate_type,
        selected_sido=sido,
        selected_region=region,
        active_tab=active_tab,
        include_sub_sigungu=include_sub_sigungu,
        menu_items=menu_items,
        current_url='/01_population/edu_dash2',

        # 차트 이미지
        age_chart_img=age_chart_img,
        region_chart_img=region_chart_img,
        age_by_region_chart_img=age_by_region_chart_img,

        # 테이블 HTML
        age_table_html=age_table_html,
        region_table_html=region_table_html,
        age_by_region_table_html=age_by_region_table_html,

        # 동적 지표 탭
        indicator_tabs=indicator_tabs,

        # 데이터
        age_data=age_data,
        age_by_region_data=age_by_region_data,
        is_initial_load=False
    )


# =============================================================================
# 메인 렌더 함수
# =============================================================================

def render(request_args):
    """메인 렌더 함수"""
    api_type = request_args.get('api_type')

    if api_type:
        result = handle_api_request(api_type, request_args)

        # download API는 send_file 객체를 직접 반환
        if api_type == 'download' and hasattr(result, 'headers'):
            return result

        return Response(
            json.dumps(result, ensure_ascii=False, default=str),
            mimetype='application/json'
        )

    # Jinja2 템플릿을 사용한 HTML 생성
    return Response(generate_dashboard_html(request_args), mimetype='text/html')
