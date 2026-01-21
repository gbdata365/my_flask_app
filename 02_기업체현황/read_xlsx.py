# -*- coding: utf-8 -*-
"""
기업체현황 엑셀 데이터 처리 및 DB 적재 스크립트
==============================================

엑셀 파일의 병합된 셀을 해제하고, dim_admin_area와 매칭하여
시군구 코드를 추가한 후 PostgreSQL에 저장합니다.

주요 기능:
1. 엑셀 병합 셀 해제 및 첫 번째 값으로 채우기
2. dim_admin_area 테이블과 매칭하여 sigungu_code 추가
3. 하이브리드 테이블 구조로 DB 적재 (giup_summary + 상세 테이블)
4. '*' 값(통계적 비밀보호)은 1로 변환

사용법:
    python read_xlsx.py                     # 모든 파일 처리 및 DB 적재
    python read_xlsx.py --file data/파일명.xlsx  # 단일 파일 처리
    python read_xlsx.py --init              # 테이블 초기화만
"""

import os
import re
import sys
import argparse
import shutil
from pathlib import Path
from typing import Optional, Tuple, Dict

import pandas as pd
import numpy as np
from loguru import logger
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

# 상위 디렉토리의 module 추가
sys.path.insert(0, str(Path(__file__).parent.parent))

from module.db import get_db_connection


# =============================================================================
# 설정
# =============================================================================
DATA_DIR = Path(__file__).parent / "data"


# =============================================================================
# 엑셀 병합 해제 기능
# =============================================================================
def unmerge_and_fill(filepath: Path, output_path: Path = None) -> int:
    """
    엑셀 파일의 모든 시트에서 병합된 셀을 해제하고 첫 번째 값으로 채움

    Args:
        filepath: 원본 엑셀 파일 경로
        output_path: 출력 파일 경로 (None이면 원본 덮어쓰기)

    Returns:
        해제된 병합 영역 수
    """
    logger.info(f"병합 해제 처리: {filepath.name}")

    wb = load_workbook(filepath)
    total_merged = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        merged_ranges = list(ws.merged_cells.ranges)

        if not merged_ranges:
            continue

        logger.debug(f"  - {sheet_name}: {len(merged_ranges)}개 병합 영역")

        for merged_range in merged_ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            first_value = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(str(merged_range))

            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row=row, column=col).value = first_value

            total_merged += 1

    save_path = output_path or filepath
    wb.save(save_path)
    wb.close()

    logger.info(f"  → 병합 해제 완료: {total_merged}개")
    return total_merged


# =============================================================================
# 기준시기 파싱
# =============================================================================
def parse_base_ym(raw_value: str) -> Tuple[str, str]:
    """
    원본 기준시기 → (표준화된 base_ym, data_type) 변환

    예시:
        '2023년' → ('202312', 'annual')
        '2024년 1분기' → ('202403', 'quarterly')
        '2025년 10월' → ('202510', 'monthly')
    """
    raw_value = str(raw_value).strip()

    # 연간: "2023년"
    match = re.match(r'(\d{4})년$', raw_value)
    if match:
        year = match.group(1)
        return f"{year}12", "annual"

    # 분기: "2024년 1분기"
    match = re.match(r'(\d{4})년\s*(\d)분기', raw_value)
    if match:
        year = match.group(1)
        quarter = int(match.group(2))
        month = str(quarter * 3).zfill(2)
        return f"{year}{month}", "quarterly"

    # 월간: "2025년 10월"
    match = re.match(r'(\d{4})년\s*(\d{1,2})월', raw_value)
    if match:
        year = match.group(1)
        month = match.group(2).zfill(2)
        return f"{year}{month}", "monthly"

    logger.warning(f"파싱 실패: {raw_value}")
    return raw_value, "unknown"


def clean_value(val, convert_star_to: int = 1):
    """
    값 정리
    - '*' → convert_star_to (기본 1, 3 이하 비식별화)
    - NaN → None
    - 숫자 변환
    """
    if pd.isna(val):
        return None
    if val == '*':
        return convert_star_to
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return val


# =============================================================================
# 테이블 생성 SQL
# =============================================================================
CREATE_TABLES_SQL = """
-- 기존 테이블 삭제
DROP TABLE IF EXISTS giup_detail_stats CASCADE;
DROP TABLE IF EXISTS giup_detail_corp_size CASCADE;
DROP TABLE IF EXISTS giup_detail_main_biz CASCADE;
DROP TABLE IF EXISTS giup_detail_age_group CASCADE;
DROP TABLE IF EXISTS giup_detail_industry CASCADE;
DROP TABLE IF EXISTS giup_detail_status CASCADE;
DROP TABLE IF EXISTS giup_detail_gender CASCADE;
DROP TABLE IF EXISTS giup_detail_org_type CASCADE;
DROP TABLE IF EXISTS giup_summary CASCADE;

-- 메인 테이블: 공통 정보
CREATE TABLE giup_summary (
    id SERIAL PRIMARY KEY,
    base_ym VARCHAR(6) NOT NULL,           -- 표준화된 기준년월 (JOIN용): 202312, 202403 등
    base_ym1 VARCHAR(30) NOT NULL,         -- 원본 기준시기: 2023년, 2024년 1분기
    data_type VARCHAR(15) NOT NULL,        -- annual, quarterly, monthly
    sido_nm VARCHAR(30),                   -- 시도명
    sigun_nm VARCHAR(30),                  -- 시군구명
    sigun_cd VARCHAR(5),                   -- 시군구코드 5자리 (dim_admin_area 매칭)
    created_at TIMESTAMP DEFAULT NOW(),
    UNIQUE (base_ym, data_type, sido_nm, sigun_nm)
);

CREATE INDEX idx_giup_summary_base_ym ON giup_summary(base_ym);
CREATE INDEX idx_giup_summary_sigun_cd ON giup_summary(sigun_cd);
CREATE INDEX idx_giup_summary_data_type ON giup_summary(data_type);

-- 조직형태별
CREATE TABLE giup_detail_org_type (
    id SERIAL PRIMARY KEY,
    summary_id INTEGER NOT NULL REFERENCES giup_summary(id) ON DELETE CASCADE,
    indiv_biz INTEGER,           -- 개인사업체
    corp INTEGER,                -- 회사법인
    corp_other INTEGER,          -- 회사이외법인
    non_corp INTEGER,            -- 비법인단체
    gov_local INTEGER,           -- 국가지방자치단체
    total INTEGER                -- 합계
);

-- 대표자성별별
CREATE TABLE giup_detail_gender (
    id SERIAL PRIMARY KEY,
    summary_id INTEGER NOT NULL REFERENCES giup_summary(id) ON DELETE CASCADE,
    blank INTEGER,               -- (공백)
    male INTEGER,                -- 남자
    female INTEGER,              -- 여자
    total INTEGER                -- 합계
);

-- 폐업여부별
CREATE TABLE giup_detail_status (
    id SERIAL PRIMARY KEY,
    summary_id INTEGER NOT NULL REFERENCES giup_summary(id) ON DELETE CASCADE,
    active INTEGER,              -- 영업중
    closed INTEGER,              -- 폐업
    total INTEGER                -- 합계
);

-- 산업분류별 (대분류)
CREATE TABLE giup_detail_industry (
    id SERIAL PRIMARY KEY,
    summary_id INTEGER NOT NULL REFERENCES giup_summary(id) ON DELETE CASCADE,
    blank INTEGER,               -- (공백)
    ind_a INTEGER,               -- 농업,임업,어업
    ind_b INTEGER,               -- 광업
    ind_c INTEGER,               -- 제조업
    ind_d INTEGER,               -- 전기가스공급업
    ind_e INTEGER,               -- 수도하수폐기물
    ind_f INTEGER,               -- 건설업
    ind_g INTEGER,               -- 도매및소매업
    ind_h INTEGER,               -- 운수및창고업
    ind_i INTEGER,               -- 숙박및음식점업
    ind_j INTEGER,               -- 정보통신업
    ind_k INTEGER,               -- 금융및보험업
    ind_l INTEGER,               -- 부동산업
    ind_m INTEGER,               -- 전문과학기술서비스
    ind_n INTEGER,               -- 사업시설관리
    ind_o INTEGER,               -- 공공행정
    ind_p INTEGER,               -- 교육서비스업
    ind_q INTEGER,               -- 보건사회복지
    ind_r INTEGER,               -- 예술스포츠여가
    ind_s INTEGER,               -- 협회및개인서비스
    ind_t INTEGER,               -- 가구내고용활동
    ind_u INTEGER,               -- 국제외국기관
    total INTEGER                -- 합계
);

-- 연령그룹별 (연간/월간만)
CREATE TABLE giup_detail_age_group (
    id SERIAL PRIMARY KEY,
    summary_id INTEGER NOT NULL REFERENCES giup_summary(id) ON DELETE CASCADE,
    age_under_19 INTEGER,        -- ~19세이하 / ~19세
    age_20_early INTEGER,        -- 20대초
    age_20_late INTEGER,         -- 20대후
    age_30_early INTEGER,        -- 30대초
    age_30_late INTEGER,         -- 30대후
    age_40_early INTEGER,        -- 40대초
    age_40_late INTEGER,         -- 40대후
    age_50_early INTEGER,        -- 50대초
    age_50_late INTEGER,         -- 50대후
    age_60_early INTEGER,        -- 60대초
    age_60_late INTEGER,         -- 60대후
    age_70_early INTEGER,        -- 70대초
    age_70_late INTEGER,         -- 70대후
    age_80_early INTEGER,        -- 80대초
    age_80_over INTEGER,         -- 80대후이상
    blank INTEGER,               -- (공백)
    total INTEGER                -- 합계
);

-- 대표사업체별
CREATE TABLE giup_detail_main_biz (
    id SERIAL PRIMARY KEY,
    summary_id INTEGER NOT NULL REFERENCES giup_summary(id) ON DELETE CASCADE,
    blank INTEGER,               -- (공백)
    total INTEGER                -- 합계
);

-- 기업규모별 (연간만)
CREATE TABLE giup_detail_corp_size (
    id SERIAL PRIMARY KEY,
    summary_id INTEGER NOT NULL REFERENCES giup_summary(id) ON DELETE CASCADE,
    blank INTEGER,               -- (공백)
    large_other INTEGER,         -- 기타대기업
    mid_large INTEGER,           -- 중견기업
    mid INTEGER,                 -- 중기업
    small INTEGER,               -- 소기업
    micro INTEGER,               -- 소상공인
    excluded INTEGER,            -- 기업규모판정제외사업자
    sangchul INTEGER,            -- 상출기업
    total INTEGER                -- 합계
);

-- 수치형통계
CREATE TABLE giup_detail_stats (
    id SERIAL PRIMARY KEY,
    summary_id INTEGER NOT NULL REFERENCES giup_summary(id) ON DELETE CASCADE,
    emp_count INTEGER,           -- 기업종사자수_건수
    emp_sum BIGINT,              -- 기업종사자수_합계
    emp_avg NUMERIC(15,2),       -- 기업종사자수_평균
    emp_blank INTEGER,           -- 기업종사자수_공백건수
    sales_count INTEGER,         -- 기업매출금액_건수
    sales_sum BIGINT,            -- 기업매출금액_합계
    sales_avg NUMERIC(20,2),     -- 기업매출금액_평균
    sales_blank INTEGER,         -- 기업매출금액_공백건수
    regular_count INTEGER,       -- 기업상용근로자수_건수
    regular_sum BIGINT,          -- 기업상용근로자수_합계
    regular_avg NUMERIC(15,2),   -- 기업상용근로자수_평균
    regular_blank INTEGER,       -- 기업상용근로자수_공백건수
    temp_count INTEGER,          -- 기업임시일용근로자수_건수
    temp_sum BIGINT,             -- 기업임시일용근로자수_합계
    temp_avg NUMERIC(15,2),      -- 기업임시일용근로자수_평균
    temp_blank INTEGER           -- 기업임시일용근로자수_공백건수
);
"""


# =============================================================================
# 시군구 코드 매핑 (dim_admin_area 테이블 사용)
# =============================================================================
def get_sigun_code_map(conn) -> dict:
    """
    시도명+시군구명 → sigungu_code 매핑 딕셔너리 생성
    dim_admin_area 테이블에서 행정표준코드 5자리(sigungu_code) 사용
    """
    cursor = conn.cursor()
    cursor.execute("""
        SELECT DISTINCT sido_nm, sigungu_nm, sigungu_code
        FROM dim_admin_area
        WHERE sido_nm IS NOT NULL AND sigungu_nm IS NOT NULL AND sigungu_code IS NOT NULL
    """)
    result = {}
    for row in cursor.fetchall():
        key = (row[0], row[1])
        result[key] = row[2]
    cursor.close()
    return result


# =============================================================================
# 컬럼 매핑 설정 (한글 → 영문) - 레거시 호환용
# =============================================================================

# 기본 컬럼 매핑
BASE_COLUMN_MAP: Dict[str, str] = {
    '기준시기': 'period_str',
    '시도명': 'sido_nm',
    '시군구명': 'sigungu_nm',
}

# 조직형태별 컬럼
ORG_TYPE_MAP: Dict[str, str] = {
    '개인사업체': 'org_individual',
    '국가지방자치단체': 'org_government',
    '비법인단체': 'org_unincorporated',
    '회사법인': 'org_corporation',
    '회사이외법인': 'org_nonprofit_corp',
    '합계': 'org_total',
}

# 대표자성별별 컬럼
GENDER_MAP: Dict[str, str] = {
    '(공백)': 'gender_unknown',
    'blank': 'gender_unknown',
    '남자': 'gender_male',
    '여자': 'gender_female',
    '합계': 'gender_total',
}

# 폐업여부별 컬럼
STATUS_MAP: Dict[str, str] = {
    '영업중': 'status_active',
    '폐업': 'status_closed',
    '합계': 'status_total',
}

# 산업분류별 컬럼 (한국표준산업분류)
INDUSTRY_MAP: Dict[str, str] = {
    '(공백)': 'ind_unknown',
    'blank': 'ind_unknown',
    '농업,임업,어업': 'ind_agriculture',
    '농업_임업_어업': 'ind_agriculture',
    '광업': 'ind_mining',
    '제조업': 'ind_manufacturing',
    '전기가스공급업': 'ind_utilities',
    '수도하수폐기물': 'ind_water_waste',
    '건설업': 'ind_construction',
    '도매및소매업': 'ind_wholesale_retail',
    '운수및창고업': 'ind_transportation',
    '숙박및음식점업': 'ind_accommodation_food',
    '정보통신업': 'ind_ict',
    '금융및보험업': 'ind_finance',
    '부동산업': 'ind_real_estate',
    '전문과학기술서비스': 'ind_professional',
    '사업시설관리': 'ind_admin_support',
    '공공행정': 'ind_public_admin',
    '교육서비스업': 'ind_education',
    '보건사회복지': 'ind_health_welfare',
    '예술스포츠여가': 'ind_arts_sports',
    '협회및개인서비스': 'ind_other_services',
    '가구내고용활동': 'ind_household',
    '국제외국기관': 'ind_international',
    '합계': 'ind_total',
}

# 대표사업체별 컬럼
MAIN_BIZ_MAP: Dict[str, str] = {
    '(공백)': 'main_biz_unknown',
    'blank': 'main_biz_unknown',
    '합계': 'main_biz_total',
}

# 수치형통계 컬럼
STATS_MAP: Dict[str, str] = {
    '기업종사자수_건수': 'emp_count',
    '기업종사자수_합계': 'emp_total',
    '기업종사자수_평균': 'emp_avg',
    '기업종사자수_공백건수': 'emp_null_count',
    '기업매출금액_건수': 'revenue_count',
    '기업매출금액_합계': 'revenue_total',
    '기업매출금액_평균': 'revenue_avg',
    '기업매출금액_공백건수': 'revenue_null_count',
    '기업상용근로자수_건수': 'regular_emp_count',
    '기업상용근로자수_합계': 'regular_emp_total',
    '기업상용근로자수_평균': 'regular_emp_avg',
    '기업상용근로자수_공백건수': 'regular_emp_null_count',
    '기업임시일용근로자수_건수': 'temp_emp_count',
    '기업임시일용근로자수_합계': 'temp_emp_total',
    '기업임시일용근로자수_평균': 'temp_emp_avg',
    '기업임시일용근로자수_공백건수': 'temp_emp_null_count',
}

# 시트별 접두사
SHEET_PREFIX_MAP: Dict[str, str] = {
    '조직형태별': 'org',
    '대표자성별별': 'gender',
    '폐업여부별': 'status',
    '산업분류별': 'ind',
    '대표사업체별': 'main_biz',
    '수치형통계': 'stats',
}

# 영문 → 한글 역매핑 (표시용)
DISPLAY_NAME_MAP: Dict[str, str] = {
    # 기본 컬럼
    'base_ym': '기준년월',
    'year': '연도',
    'period_type': '구분',
    'period_str': '기준시기',
    'sido_nm': '시도명',
    'sigungu_nm': '시군구명',

    # 조직형태별
    'org_individual': '개인사업체',
    'org_government': '국가지방자치단체',
    'org_unincorporated': '비법인단체',
    'org_corporation': '회사법인',
    'org_nonprofit_corp': '회사이외법인',
    'org_total': '조직형태_합계',

    # 대표자성별별
    'gender_unknown': '성별미상',
    'gender_male': '남성대표',
    'gender_female': '여성대표',
    'gender_total': '성별_합계',

    # 폐업여부별
    'status_active': '영업중',
    'status_closed': '폐업',
    'status_total': '영업상태_합계',

    # 산업분류별
    'ind_unknown': '산업미분류',
    'ind_agriculture': '농림어업',
    'ind_mining': '광업',
    'ind_manufacturing': '제조업',
    'ind_utilities': '전기가스공급',
    'ind_water_waste': '수도하수폐기물',
    'ind_construction': '건설업',
    'ind_wholesale_retail': '도소매업',
    'ind_transportation': '운수창고업',
    'ind_accommodation_food': '숙박음식점',
    'ind_ict': '정보통신업',
    'ind_finance': '금융보험업',
    'ind_real_estate': '부동산업',
    'ind_professional': '전문과학기술',
    'ind_admin_support': '사업시설관리',
    'ind_public_admin': '공공행정',
    'ind_education': '교육서비스',
    'ind_health_welfare': '보건복지',
    'ind_arts_sports': '예술스포츠여가',
    'ind_other_services': '협회개인서비스',
    'ind_household': '가구내고용',
    'ind_international': '국제기관',
    'ind_total': '산업_합계',

    # 대표사업체별
    'main_biz_unknown': '대표사업체미상',
    'main_biz_total': '대표사업체_합계',

    # 수치형통계
    'emp_count': '종사자수_건수',
    'emp_total': '종사자수_합계',
    'emp_avg': '종사자수_평균',
    'emp_null_count': '종사자수_공백',
    'revenue_count': '매출금액_건수',
    'revenue_total': '매출금액_합계',
    'revenue_avg': '매출금액_평균',
    'revenue_null_count': '매출금액_공백',
    'regular_emp_count': '상용근로자_건수',
    'regular_emp_total': '상용근로자_합계',
    'regular_emp_avg': '상용근로자_평균',
    'regular_emp_null_count': '상용근로자_공백',
    'temp_emp_count': '임시일용_건수',
    'temp_emp_total': '임시일용_합계',
    'temp_emp_avg': '임시일용_평균',
    'temp_emp_null_count': '임시일용_공백',
}


# =============================================================================
# 데이터 로더 클래스
# =============================================================================
class GiupDataLoader:
    """기업체현황 데이터 로더 - 병합해제, dim_admin_area 매칭, DB 적재"""

    def __init__(self):
        self.conn = get_db_connection()
        self.sigun_map = {}

    def init_tables(self, drop_existing: bool = True):
        """테이블 생성 (재적재 시 기존 테이블 삭제)"""
        cursor = self.conn.cursor()

        if drop_existing:
            logger.info("기존 테이블 삭제 및 재생성...")
            cursor.execute(CREATE_TABLES_SQL)

        self.conn.commit()
        cursor.close()

        # 시군구 코드 매핑 로드 (dim_admin_area 테이블 사용)
        self.sigun_map = get_sigun_code_map(self.conn)
        logger.info(f"시군구 코드 매핑 로드 완료 (dim_admin_area): {len(self.sigun_map)}건")

    def get_sigun_cd(self, sido_nm: str, sigun_nm: str) -> Optional[str]:
        """시도명+시군구명으로 시군구코드 조회"""
        return self.sigun_map.get((sido_nm, sigun_nm))

    def load_excel_file(self, filepath: Path) -> dict:
        """엑셀 파일 로드 및 파싱"""
        logger.info(f"파일 로드: {filepath.name}")
        xl = pd.ExcelFile(filepath)

        sheets = {}
        for sheet_name in xl.sheet_names:
            if sheet_name == '결측치현황':
                continue  # 결측치현황은 제외
            df = pd.read_excel(filepath, sheet_name=sheet_name)
            sheets[sheet_name] = df
            logger.debug(f"  - {sheet_name}: {len(df)}행")

        return sheets

    def process_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """DataFrame 전처리"""
        df = df.copy()

        # 시도명, 시군구명 forward fill
        if '시도명' in df.columns:
            df['시도명'] = df['시도명'].ffill()
        if '시군구명' in df.columns:
            df['시군구명'] = df['시군구명'].ffill()

        # 기준시기 forward fill
        if '기준시기' in df.columns:
            df['기준시기'] = df['기준시기'].ffill()
            df['base_ym'] = ''
            df['data_type'] = ''
            for idx, row in df.iterrows():
                base_ym, data_type = parse_base_ym(row['기준시기'])
                df.at[idx, 'base_ym'] = base_ym
                df.at[idx, 'data_type'] = data_type

        return df

    def insert_summary(self, cursor, base_ym: str, base_ym1: str, data_type: str,
                       sido_nm: str, sigun_nm: str) -> int:
        """메인 테이블에 INSERT 후 id 반환"""
        sigun_cd = self.get_sigun_cd(sido_nm, sigun_nm)

        cursor.execute("""
            INSERT INTO giup_summary (base_ym, base_ym1, data_type, sido_nm, sigun_nm, sigun_cd)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON CONFLICT (base_ym, data_type, sido_nm, sigun_nm) DO UPDATE
            SET sigun_cd = EXCLUDED.sigun_cd
            RETURNING id
        """, (base_ym, base_ym1, data_type, sido_nm, sigun_nm, sigun_cd))

        return cursor.fetchone()[0]

    def insert_org_type(self, cursor, summary_id: int, row: pd.Series):
        """조직형태별 INSERT"""
        cursor.execute("""
            INSERT INTO giup_detail_org_type
            (summary_id, indiv_biz, corp, corp_other, non_corp, gov_local, total)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (
            summary_id,
            clean_value(row.get('개인사업체')),
            clean_value(row.get('회사법인')),
            clean_value(row.get('회사이외법인')),
            clean_value(row.get('비법인단체')),
            clean_value(row.get('국가지방자치단체')),
            clean_value(row.get('합계'))
        ))

    def insert_gender(self, cursor, summary_id: int, row: pd.Series):
        """대표자성별별 INSERT"""
        cursor.execute("""
            INSERT INTO giup_detail_gender
            (summary_id, blank, male, female, total)
            VALUES (%s, %s, %s, %s, %s)
        """, (
            summary_id,
            clean_value(row.get('(공백)')),
            clean_value(row.get('남자')),
            clean_value(row.get('여자')),
            clean_value(row.get('합계'))
        ))

    def insert_status(self, cursor, summary_id: int, row: pd.Series):
        """폐업여부별 INSERT"""
        cursor.execute("""
            INSERT INTO giup_detail_status
            (summary_id, active, closed, total)
            VALUES (%s, %s, %s, %s)
        """, (
            summary_id,
            clean_value(row.get('영업중')),
            clean_value(row.get('폐업')),
            clean_value(row.get('합계'))
        ))

    def insert_industry(self, cursor, summary_id: int, row: pd.Series):
        """산업분류별 INSERT"""
        cursor.execute("""
            INSERT INTO giup_detail_industry
            (summary_id, blank, ind_a, ind_b, ind_c, ind_d, ind_e, ind_f, ind_g, ind_h, ind_i,
             ind_j, ind_k, ind_l, ind_m, ind_n, ind_o, ind_p, ind_q, ind_r, ind_s, ind_t, ind_u, total)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            summary_id,
            clean_value(row.get('(공백)')),
            clean_value(row.get('농업,임업,어업')),
            clean_value(row.get('광업')),
            clean_value(row.get('제조업')),
            clean_value(row.get('전기가스공급업')),
            clean_value(row.get('수도하수폐기물')),
            clean_value(row.get('건설업')),
            clean_value(row.get('도매및소매업')),
            clean_value(row.get('운수및창고업')),
            clean_value(row.get('숙박및음식점업')),
            clean_value(row.get('정보통신업')),
            clean_value(row.get('금융및보험업')),
            clean_value(row.get('부동산업')),
            clean_value(row.get('전문과학기술서비스')),
            clean_value(row.get('사업시설관리')),
            clean_value(row.get('공공행정')),
            clean_value(row.get('교육서비스업')),
            clean_value(row.get('보건사회복지')),
            clean_value(row.get('예술스포츠여가')),
            clean_value(row.get('협회및개인서비스')),
            clean_value(row.get('가구내고용활동')),
            clean_value(row.get('국제외국기관')),
            clean_value(row.get('합계'))
        ))

    def insert_age_group(self, cursor, summary_id: int, row: pd.Series):
        """연령그룹별 INSERT"""
        age_under_19 = row.get('~19세이하') if '~19세이하' in row.index else row.get('~19세')

        # 엑셀 헤더가 '20대말', '30대말' 등으로 되어 있음 (초/말 5세 단위)
        cursor.execute("""
            INSERT INTO giup_detail_age_group
            (summary_id, age_under_19, age_20_early, age_20_late, age_30_early, age_30_late,
             age_40_early, age_40_late, age_50_early, age_50_late, age_60_early, age_60_late,
             age_70_early, age_70_late, age_80_early, age_80_over, blank, total)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            summary_id,
            clean_value(age_under_19),
            clean_value(row.get('20대초')),
            clean_value(row.get('20대말')),  # 엑셀: 20대말
            clean_value(row.get('30대초')),
            clean_value(row.get('30대말')),  # 엑셀: 30대말
            clean_value(row.get('40대초')),
            clean_value(row.get('40대말')),  # 엑셀: 40대말
            clean_value(row.get('50대초')),
            clean_value(row.get('50대말')),  # 엑셀: 50대말
            clean_value(row.get('60대초')),
            clean_value(row.get('60대말')),  # 엑셀: 60대말
            clean_value(row.get('70대초')),
            clean_value(row.get('70대말')),  # 엑셀: 70대말
            clean_value(row.get('80대이상')),  # 엑셀: 80대이상
            clean_value(row.get('80대이상')),  # 중복 (원본에 80대초 없음)
            clean_value(row.get('(공백)')),
            clean_value(row.get('합계'))
        ))

    def insert_main_biz(self, cursor, summary_id: int, row: pd.Series):
        """대표사업체별 INSERT"""
        cursor.execute("""
            INSERT INTO giup_detail_main_biz
            (summary_id, blank, total)
            VALUES (%s, %s, %s)
        """, (
            summary_id,
            clean_value(row.get('(공백)')),
            clean_value(row.get('합계'))
        ))

    def insert_corp_size(self, cursor, summary_id: int, row: pd.Series):
        """기업규모별 INSERT"""
        cursor.execute("""
            INSERT INTO giup_detail_corp_size
            (summary_id, blank, large_other, mid_large, mid, small, micro, excluded, sangchul, total)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            summary_id,
            clean_value(row.get('(공백)')),
            clean_value(row.get('기타대기업')),
            clean_value(row.get('중견기업')),
            clean_value(row.get('중기업')),
            clean_value(row.get('소기업')),
            clean_value(row.get('소상공인')),
            clean_value(row.get('기업규모판정제외사업자')),
            clean_value(row.get('상출기업')),
            clean_value(row.get('합계'))
        ))

    def insert_stats(self, cursor, summary_id: int, row: pd.Series):
        """수치형통계 INSERT"""
        def clean_numeric(val):
            if pd.isna(val) or val == '*':
                return None
            try:
                return float(val)
            except:
                return None

        cursor.execute("""
            INSERT INTO giup_detail_stats
            (summary_id, emp_count, emp_sum, emp_avg, emp_blank,
             sales_count, sales_sum, sales_avg, sales_blank,
             regular_count, regular_sum, regular_avg, regular_blank,
             temp_count, temp_sum, temp_avg, temp_blank)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            summary_id,
            clean_value(row.get('기업종사자수_건수')),
            clean_value(row.get('기업종사자수_합계')),
            clean_numeric(row.get('기업종사자수_평균')),
            clean_value(row.get('기업종사자수_공백건수')),
            clean_value(row.get('기업매출금액_건수')),
            clean_value(row.get('기업매출금액_합계')),
            clean_numeric(row.get('기업매출금액_평균')),
            clean_value(row.get('기업매출금액_공백건수')),
            clean_value(row.get('기업상용근로자수_건수')),
            clean_value(row.get('기업상용근로자수_합계')),
            clean_numeric(row.get('기업상용근로자수_평균')),
            clean_value(row.get('기업상용근로자수_공백건수')),
            clean_value(row.get('기업임시일용근로자수_건수')),
            clean_value(row.get('기업임시일용근로자수_합계')),
            clean_numeric(row.get('기업임시일용근로자수_평균')),
            clean_value(row.get('기업임시일용근로자수_공백건수'))
        ))

    def load_file(self, filepath: Path, backup: bool = True):
        """단일 파일 로드 및 DB 적재 (병합 해제 포함)"""

        # 1. 병합 해제 처리
        if backup:
            backup_path = filepath.with_suffix('.xlsx.bak')
            if not backup_path.exists():
                shutil.copy(filepath, backup_path)
                logger.info(f"백업 생성: {backup_path.name}")

        unmerge_and_fill(filepath)

        # 2. 엑셀 파일 로드
        sheets = self.load_excel_file(filepath)
        cursor = self.conn.cursor()

        # 조직형태별 시트를 기준으로 summary 레코드 생성
        if '조직형태별' not in sheets:
            logger.error(f"조직형태별 시트가 없습니다: {filepath.name}")
            return

        org_df = self.process_dataframe(sheets['조직형태별'])

        # summary_id 캐시 (base_ym, data_type, sido_nm, sigun_nm) → summary_id
        summary_cache = {}

        for idx, row in org_df.iterrows():
            if pd.isna(row.get('시도명')) or pd.isna(row.get('시군구명')):
                continue

            key = (row['base_ym'], row['data_type'], row['시도명'], row['시군구명'])

            # summary INSERT
            summary_id = self.insert_summary(
                cursor,
                row['base_ym'],
                str(row['기준시기']),
                row['data_type'],
                row['시도명'],
                row['시군구명']
            )
            summary_cache[key] = summary_id

            # 조직형태별 INSERT
            self.insert_org_type(cursor, summary_id, row)

        logger.info(f"  - giup_summary: {len(summary_cache)}건")

        # 각 시트별 처리
        sheet_handlers = {
            '대표자성별별': self.insert_gender,
            '폐업여부별': self.insert_status,
            '산업분류별': self.insert_industry,
            '연령그룹별': self.insert_age_group,
            '대표사업체별': self.insert_main_biz,
            '기업규모별': self.insert_corp_size,
            '수치형통계': self.insert_stats,
        }

        for sheet_name, handler in sheet_handlers.items():
            if sheet_name not in sheets:
                continue

            df = self.process_dataframe(sheets[sheet_name])
            count = 0

            for idx, row in df.iterrows():
                if pd.isna(row.get('시도명')) or pd.isna(row.get('시군구명')):
                    continue

                key = (row['base_ym'], row['data_type'], row['시도명'], row['시군구명'])
                summary_id = summary_cache.get(key)

                if summary_id:
                    handler(cursor, summary_id, row)
                    count += 1

            logger.info(f"  - {sheet_name}: {count}건")

        self.conn.commit()
        cursor.close()

    def load_all_files(self, backup: bool = True):
        """data 폴더의 모든 집계표 파일 로드"""
        files = sorted(DATA_DIR.glob("(수정)집계표_*.xlsx"))

        logger.info(f"총 {len(files)}개 파일 발견")

        for filepath in files:
            self.load_file(filepath, backup=backup)

        logger.info("모든 파일 로드 완료")

    def get_summary(self):
        """적재 결과 요약"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM giup_summary")
        summary_count = cursor.fetchone()[0]

        cursor.execute("SELECT data_type, COUNT(*) FROM giup_summary GROUP BY data_type ORDER BY data_type")
        type_counts = cursor.fetchall()

        cursor.execute("""
            SELECT COUNT(*) FROM giup_summary WHERE sigun_cd IS NOT NULL
        """)
        matched_count = cursor.fetchone()[0]

        cursor.close()
        return summary_count, type_counts, matched_count

    def close(self):
        """연결 종료"""
        self.conn.close()


# =============================================================================
# 메인 함수
# =============================================================================
def main():
    """메인 실행"""
    parser = argparse.ArgumentParser(
        description='기업체현황 엑셀 데이터 처리 및 DB 적재',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
예시:
  python read_xlsx.py                           # 모든 파일 처리 및 DB 적재
  python read_xlsx.py --file data/파일명.xlsx   # 단일 파일만 처리
  python read_xlsx.py --init                    # 테이블 초기화만

기능:
  1. 엑셀 병합 셀 해제 및 첫 번째 값으로 채우기
  2. dim_admin_area 테이블과 매칭하여 sigungu_code 추가
  3. 하이브리드 테이블 구조로 DB 적재 (giup_summary + 상세 테이블)
        """)
    parser.add_argument('--file', '-f', help='단일 엑셀 파일 경로')
    parser.add_argument('--init', action='store_true', help='테이블 초기화만 (데이터 적재 안함)')
    parser.add_argument('--no-backup', action='store_true', help='백업 파일 생성 안함')

    args = parser.parse_args()
    os.chdir(Path(__file__).parent)

    logger.info("=" * 60)
    logger.info("기업체현황 집계표 DB 적재 시작")
    logger.info("=" * 60)

    loader = GiupDataLoader()

    try:
        # 테이블 초기화 (기존 데이터 삭제 후 재생성)
        loader.init_tables(drop_existing=True)

        if args.init:
            logger.info("테이블 초기화 완료 (--init 옵션)")
            return

        backup = not args.no_backup

        if args.file:
            # 단일 파일 처리
            filepath = Path(args.file)
            if not filepath.exists():
                logger.error(f"파일 없음: {args.file}")
                return
            loader.load_file(filepath, backup=backup)
        else:
            # 모든 파일 처리
            loader.load_all_files(backup=backup)

        # 결과 확인
        summary_count, type_counts, matched_count = loader.get_summary()

        logger.info("=" * 60)
        logger.info(f"적재 완료: giup_summary 총 {summary_count}건")
        for dt, cnt in type_counts:
            logger.info(f"  - {dt}: {cnt}건")
        logger.info(f"dim_admin_area 매칭: {matched_count}건 ({matched_count/summary_count*100:.1f}%)")
        logger.info("=" * 60)

    except Exception as e:
        logger.error(f"오류 발생: {e}")
        raise
    finally:
        loader.close()


if __name__ == '__main__':
    main()


