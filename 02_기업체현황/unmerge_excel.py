# -*- coding: utf-8 -*-
"""
엑셀 파일의 병합된 셀을 해제하고 첫 번째 값으로 채우는 스크립트
"""

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from loguru import logger
import shutil


def unmerge_and_fill(filepath: Path, output_path: Path = None):
    """
    엑셀 파일의 모든 시트에서 병합된 셀을 해제하고 첫 번째 값으로 채움

    Args:
        filepath: 원본 엑셀 파일 경로
        output_path: 출력 파일 경로 (None이면 원본 덮어쓰기)
    """
    logger.info(f"파일 처리 시작: {filepath.name}")

    # 워크북 로드
    wb = load_workbook(filepath)

    total_merged = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 병합된 셀 목록 복사 (순회 중 수정하면 안 되므로)
        merged_ranges = list(ws.merged_cells.ranges)

        if not merged_ranges:
            continue

        logger.info(f"  - {sheet_name}: {len(merged_ranges)}개 병합 영역")

        for merged_range in merged_ranges:
            # 병합 영역의 경계 좌표 가져오기
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))

            # 병합된 셀의 첫 번째 값 가져오기
            first_value = ws.cell(row=min_row, column=min_col).value

            # 병합 해제
            ws.unmerge_cells(str(merged_range))

            # 병합되었던 모든 셀에 첫 번째 값 채우기
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row=row, column=col).value = first_value

            total_merged += 1

    # 저장
    save_path = output_path or filepath
    wb.save(save_path)
    wb.close()

    logger.info(f"  → 저장 완료: {save_path.name} (총 {total_merged}개 병합 해제)")

    return total_merged


def process_all_files(data_dir: Path, backup: bool = True):
    """
    data 폴더의 모든 집계표 파일 처리

    Args:
        data_dir: 데이터 폴더 경로
        backup: 백업 생성 여부
    """
    files = sorted(data_dir.glob("(수정)집계표_*.xlsx"))

    logger.info(f"총 {len(files)}개 파일 발견")
    logger.info("=" * 60)

    for filepath in files:
        # 백업 생성
        if backup:
            backup_path = filepath.with_suffix('.xlsx.bak')
            if not backup_path.exists():
                shutil.copy(filepath, backup_path)
                logger.info(f"백업 생성: {backup_path.name}")

        # 병합 해제 처리
        unmerge_and_fill(filepath)

    logger.info("=" * 60)
    logger.info("모든 파일 처리 완료")


if __name__ == '__main__':
    data_dir = Path(__file__).parent / "data"
    process_all_files(data_dir, backup=True)
