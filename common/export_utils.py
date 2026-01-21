"""
================================================================================
공통 내보내기 유틸리티 모듈 (export_utils.py)
================================================================================
다양한 분야의 데이터를 Excel, Markdown, HTML 형식으로 내보내기 위한 범용 모듈

사용 예시:
    from common.export_utils import DataExporter

    exporter = DataExporter(
        tables_data={'시도별_현황': data1, '지표별_현황': data2},
        charts_data={'시도별_현황': chart_base64, ...},
        sort_config={'column': 'pop_202512', 'direction': 'desc'}
    )
    exporter.export_all(output_dir, filename_base)

수정 이력:
    2025-01-21: 최초 생성 - 정렬 파라미터 지원, 탭 형식 HTML 내보내기
================================================================================
"""

from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any
import traceback


class DataExporter:
    """
    데이터 내보내기 클래스

    다양한 분야(인구, 경제, 환경 등)의 데이터를 여러 형식으로 내보내기 지원
    - Excel (.xlsx)
    - Markdown (.md)
    - HTML (.html) - 탭 형식, 차트 포함

    Attributes:
        tables_data: 테이블 데이터 딕셔너리 {시트명: {data: [...], headers: [...], ...}}
        charts_data: 차트 이미지 딕셔너리 {시트명: base64_string}
        sort_config: 정렬 설정 {column: 컬럼명, direction: 'asc'|'desc'}
        report_title: 보고서 제목
        header_labels: 헤더명 한글 매핑 딕셔너리
    """

    def __init__(
        self,
        tables_data: Dict[str, Any],
        charts_data: Optional[Dict[str, str]] = None,
        sort_config: Optional[Dict[str, str]] = None,
        report_title: str = "데이터 보고서",
        header_labels: Optional[Dict[str, str]] = None
    ):
        """
        DataExporter 초기화

        Args:
            tables_data: 내보낼 테이블 데이터
            charts_data: 차트 이미지 (base64 인코딩)
            sort_config: 정렬 설정 {'column': 컬럼명, 'direction': 'asc'|'desc'}
            report_title: 보고서 제목
            header_labels: 컬럼명 -> 한글명 매핑
        """
        self.tables_data = tables_data
        self.charts_data = charts_data or {}
        self.sort_config = sort_config
        self.report_title = report_title

        # 기본 헤더 라벨 (확장 가능)
        self.header_labels = header_labels or {
            'name': '지역',
            'sigungu_nm': '시군',
            'sido_nm': '시도',
            'age_group': '연령대',
        }

    def _sort_data(self, data: List[Dict], keep_first_row: bool = True) -> List[Dict]:
        """
        데이터 정렬 (첫 번째 행(합계)은 유지 옵션)

        Args:
            data: 정렬할 데이터 리스트
            keep_first_row: True면 첫 행(합계)을 맨 위에 유지

        Returns:
            정렬된 데이터 리스트
        """
        if not self.sort_config or not data:
            return data

        sort_col = self.sort_config.get('column', '')
        sort_dir = self.sort_config.get('direction', 'asc')

        if not sort_col:
            return data

        # 첫 번째 행(합계) 분리
        if keep_first_row and len(data) > 1:
            first_row = data[0]
            rows_to_sort = data[1:]
        else:
            first_row = None
            rows_to_sort = data

        # 정렬 수행
        try:
            def sort_key(row):
                val = row.get(sort_col, 0)
                if val is None:
                    return 0
                if isinstance(val, str):
                    return val
                return float(val) if val else 0

            reverse = (sort_dir == 'desc')
            sorted_rows = sorted(rows_to_sort, key=sort_key, reverse=reverse)

            # 첫 번째 행 복원
            if first_row:
                return [first_row] + sorted_rows
            return sorted_rows

        except Exception as e:
            print(f"정렬 오류: {e}")
            return data

    def _get_header_label(self, col_name: str) -> str:
        """
        컬럼명을 한글 라벨로 변환

        Args:
            col_name: 원본 컬럼명

        Returns:
            한글 라벨
        """
        # 사전 정의된 라벨 확인
        if col_name in self.header_labels:
            return self.header_labels[col_name]

        # 패턴 기반 변환
        if col_name.startswith('pop_'):
            return col_name.replace('pop_', '인구 ')
        elif col_name.startswith('single_'):
            return col_name.replace('single_', '1인가구 ')
        elif col_name.startswith('value_'):
            return col_name.replace('value_', '지표값 ')
        elif col_name.startswith('numerator_'):
            return col_name.replace('numerator_', '분자 ')
        elif col_name.startswith('change_'):
            return col_name.replace('change_', '증감 ')
        elif col_name.startswith('rate_') or col_name.startswith('pop_rate_'):
            return col_name.replace('rate_', '증감률 ').replace('pop_rate_', '증감률 ')
        elif col_name.startswith('elderly_rate_'):
            return col_name.replace('elderly_rate_', '고령화율 ')

        return col_name.replace('_', ' ')

    def export_to_excel(self, output_path: Path) -> bool:
        """
        Excel 파일로 내보내기 (정렬 적용)

        Args:
            output_path: 출력 파일 경로

        Returns:
            성공 여부
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

            wb = openpyxl.Workbook()

            # 스타일 정의
            header_fill = PatternFill(start_color='1243A6', end_color='1243A6', fill_type='solid')
            header_font = Font(bold=True, size=11, color='FFFFFF')
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            for idx, (sheet_name, table_data) in enumerate(self.tables_data.items()):
                # 시트 생성
                if idx == 0:
                    ws = wb.active
                    ws.title = sheet_name[:31]
                else:
                    ws = wb.create_sheet(title=sheet_name[:31])

                if not table_data.get('data'):
                    continue

                # 정렬 적용
                sorted_data = self._sort_data(table_data['data'])

                # 헤더
                start_row = 1
                if 'title' in table_data:
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
                    ws.cell(row=1, column=1, value=table_data['title'])
                    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
                    start_row = 3

                # 컬럼 헤더 (코드 컬럼 제외)
                headers = [h for h in sorted_data[0].keys()
                          if not h.endswith('_code') and h != 'sigungu_code']

                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=start_row, column=col, value=self._get_header_label(header))
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border

                # 데이터 행
                for row_idx, row_data in enumerate(sorted_data, start_row + 1):
                    for col, header in enumerate(headers, 1):
                        val = row_data.get(header, '')
                        cell = ws.cell(row=row_idx, column=col, value=val)
                        cell.border = thin_border
                        if isinstance(val, (int, float)) and col > 1:
                            cell.alignment = Alignment(horizontal='right')
                            if isinstance(val, float):
                                cell.number_format = '#,##0.00'
                            else:
                                cell.number_format = '#,##0'

            wb.save(output_path)
            return True

        except Exception as e:
            print(f"Excel 내보내기 오류: {e}")
            traceback.print_exc()
            return False

    def export_to_markdown(self, output_path: Path) -> bool:
        """
        Markdown 파일로 내보내기 (정렬 적용)

        Args:
            output_path: 출력 파일 경로

        Returns:
            성공 여부
        """
        try:
            lines = [f"# {self.report_title}\n"]
            lines.append(f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

            # 정렬 정보 표시
            if self.sort_config and self.sort_config.get('column'):
                sort_col = self._get_header_label(self.sort_config['column'])
                sort_dir = '내림차순' if self.sort_config.get('direction') == 'desc' else '오름차순'
                lines.append(f"정렬: {sort_col} ({sort_dir})\n")

            for sheet_name, table_data in self.tables_data.items():
                lines.append(f"\n## {sheet_name.replace('_', ' ')}\n")

                if 'title' in table_data:
                    lines.append(f"### {table_data['title']}\n")
                if 'description' in table_data and table_data['description']:
                    lines.append(f"*{table_data['description']}*\n")

                if not table_data.get('data'):
                    lines.append("*데이터 없음*\n")
                    continue

                # 정렬 적용
                sorted_data = self._sort_data(table_data['data'])

                # 헤더 (코드 컬럼 제외)
                headers = [h for h in sorted_data[0].keys()
                          if not h.endswith('_code') and h != 'sigungu_code']
                header_labels = [self._get_header_label(h) for h in headers]

                lines.append("| " + " | ".join(header_labels) + " |")
                lines.append("| " + " | ".join(["---"] * len(headers)) + " |")

                # 데이터 행
                for row in sorted_data:
                    values = []
                    for h in headers:
                        val = row.get(h, '')
                        if isinstance(val, float):
                            values.append(f"{val:,.2f}")
                        elif isinstance(val, int):
                            values.append(f"{val:,}")
                        else:
                            values.append(str(val))
                    lines.append("| " + " | ".join(values) + " |")

                lines.append("")

            with open(output_path, 'w', encoding='utf-8') as f:
                f.write('\n'.join(lines))

            return True

        except Exception as e:
            print(f"Markdown 내보내기 오류: {e}")
            traceback.print_exc()
            return False

    def export_to_html(self, output_path: Path) -> bool:
        """
        HTML 파일로 내보내기 (탭 형식, 차트 포함, 정렬 적용)

        Args:
            output_path: 출력 파일 경로

        Returns:
            성공 여부
        """
        try:
            # 탭 아이템 생성
            tab_items = []
            for idx, (sheet_name, table_data) in enumerate(self.tables_data.items()):
                tab_items.append({
                    'id': f"tab_{idx}",
                    'name': sheet_name.replace('_', ' '),
                    'data': table_data,
                    'chart': self.charts_data.get(sheet_name)
                })

            # 정렬 정보 텍스트
            sort_info = ""
            if self.sort_config and self.sort_config.get('column'):
                sort_col = self._get_header_label(self.sort_config['column'])
                sort_dir = '내림차순' if self.sort_config.get('direction') == 'desc' else '오름차순'
                sort_info = f"<p class='sort-info'>정렬: {sort_col} ({sort_dir})</p>"

            html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{self.report_title}</title>
    <style>
        /* ===== 기본 스타일 ===== */
        * {{ box-sizing: border-box; }}
        body {{ font-family: 'Pretendard', 'Noto Sans KR', sans-serif; margin: 0; padding: 20px; background: #f5f5f5; }}
        .container {{ max-width: 1400px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        h1 {{ color: #1243A6; border-bottom: 2px solid #1243A6; padding-bottom: 10px; margin-bottom: 20px; }}
        .meta-info {{ color: #666; margin-bottom: 10px; }}
        .sort-info {{ color: #1D64F2; font-weight: 500; margin-bottom: 20px; padding: 8px 12px; background: #e8f4fc; border-radius: 4px; display: inline-block; }}

        /* ===== 탭 스타일 ===== */
        .tab-nav {{ display: flex; flex-wrap: wrap; gap: 5px; border-bottom: 2px solid #1243A6; padding-bottom: 0; margin-bottom: 0; }}
        .tab-btn {{
            padding: 10px 20px; border: none; background: #e8f4fc; color: #1243A6;
            cursor: pointer; border-radius: 8px 8px 0 0; font-size: 14px; font-weight: 500;
            transition: all 0.2s;
        }}
        .tab-btn:hover {{ background: #c5e3f6; }}
        .tab-btn.active {{ background: #1243A6; color: white; }}
        .tab-content {{ display: none; padding: 20px 0; }}
        .tab-content.active {{ display: block; }}

        /* ===== 차트 스타일 ===== */
        .chart-container {{ text-align: center; margin: 20px 0; }}
        .chart-container img {{ max-width: 100%; height: auto; border: 1px solid #eee; border-radius: 8px; }}

        /* ===== 테이블 스타일 ===== */
        .table-container {{ overflow-x: auto; }}
        table {{ border-collapse: collapse; width: 100%; margin: 15px 0; font-size: 13px; }}
        th {{ background: #1243A6; color: white; padding: 10px 8px; text-align: center; white-space: nowrap; }}
        td {{ border: 1px solid #ddd; padding: 8px; text-align: right; }}
        td:first-child {{ text-align: left; font-weight: 500; background: #f8f9fa; }}
        tr:nth-child(even) {{ background: #f9f9f9; }}
        tr:hover {{ background: #e8f4fc; }}
        tr:first-child td {{ background: #e8f4fc; font-weight: bold; }}
        .positive {{ color: #2563eb; }}
        .negative {{ color: #dc2626; }}

        .description {{ color: #666; font-style: italic; margin: 10px 0; padding: 10px; background: #f8f9fa; border-radius: 4px; }}
        .footer {{ margin-top: 30px; padding-top: 20px; border-top: 1px solid #ddd; color: #666; font-size: 0.9em; }}
    </style>
</head>
<body>
<div class="container">
    <h1>{self.report_title}</h1>
    <p class="meta-info"><strong>생성일시:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
    {sort_info}
"""
            # 탭 네비게이션
            html += '<div class="tab-nav">'
            for idx, tab in enumerate(tab_items):
                active = ' active' if idx == 0 else ''
                html += f'<button class="tab-btn{active}" onclick="showTab(\'{tab["id"]}\')">{tab["name"]}</button>'
            html += '</div>'

            # 탭 콘텐츠
            for idx, tab in enumerate(tab_items):
                active = ' active' if idx == 0 else ''
                html += f'<div id="{tab["id"]}" class="tab-content{active}">'

                table_data = tab['data']

                # 설명
                if table_data.get('description'):
                    html += f"<p class='description'>{table_data['description']}</p>"

                # 차트
                if tab['chart']:
                    html += f'<div class="chart-container"><img src="data:image/png;base64,{tab["chart"]}" alt="{tab["name"]} 차트"></div>'

                # 테이블
                if table_data.get('data'):
                    # 정렬 적용
                    sorted_data = self._sort_data(table_data['data'])

                    # 헤더 (코드 컬럼 제외)
                    headers = [h for h in sorted_data[0].keys()
                              if not h.endswith('_code') and h != 'sigungu_code']

                    html += '<div class="table-container"><table><thead><tr>'
                    for h in headers:
                        html += f"<th>{self._get_header_label(h)}</th>"
                    html += "</tr></thead><tbody>"

                    for row in sorted_data:
                        html += "<tr>"
                        for h in headers:
                            val = row.get(h, '')
                            if isinstance(val, float):
                                css = ''
                                if 'rate' in h or 'change' in h:
                                    css = ' class="positive"' if val > 0 else ' class="negative"' if val < 0 else ''
                                html += f"<td{css}>{val:,.2f}</td>"
                            elif isinstance(val, int):
                                html += f"<td>{val:,}</td>"
                            else:
                                html += f"<td>{val}</td>"
                        html += "</tr>"

                    html += "</tbody></table></div>"
                else:
                    html += "<p><em>데이터 없음</em></p>"

                html += '</div>'

            # JavaScript
            html += """
<script>
function showTab(tabId) {
    document.querySelectorAll('.tab-content').forEach(c => c.classList.remove('active'));
    document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
    document.getElementById(tabId).classList.add('active');
    event.target.classList.add('active');
}
</script>
<div class="footer">
    <p>본 보고서는 대시보드에서 자동 생성되었습니다.</p>
</div>
</div>
</body>
</html>"""

            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html)

            return True

        except Exception as e:
            print(f"HTML 내보내기 오류: {e}")
            traceback.print_exc()
            return False

    def export_all(self, output_dir: Path, filename_base: str) -> Dict[str, Any]:
        """
        모든 형식으로 내보내기 (Excel, Markdown, HTML)

        Args:
            output_dir: 출력 디렉토리
            filename_base: 파일명 기본값 (확장자 제외)

        Returns:
            결과 딕셔너리 {'success': bool, 'files': [...], 'output_dir': str}
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        results = {'success': True, 'files': [], 'output_dir': str(output_dir)}

        # Excel
        excel_path = output_dir / f"{filename_base}.xlsx"
        if self.export_to_excel(excel_path):
            results['files'].append(str(excel_path))

        # Markdown
        md_path = output_dir / f"{filename_base}.md"
        if self.export_to_markdown(md_path):
            results['files'].append(str(md_path))

        # HTML
        html_path = output_dir / f"{filename_base}.html"
        if self.export_to_html(html_path):
            results['files'].append(str(html_path))

        results['success'] = len(results['files']) > 0
        return results


# ================================================================================
# 편의 함수들 (기존 코드 호환용)
# ================================================================================

def export_to_excel(tables_data: Dict, output_path: Path, sort_config: Dict = None) -> bool:
    """기존 코드 호환용 Excel 내보내기 함수"""
    exporter = DataExporter(tables_data, sort_config=sort_config)
    return exporter.export_to_excel(output_path)


def export_to_markdown(tables_data: Dict, output_path: Path, sort_config: Dict = None) -> bool:
    """기존 코드 호환용 Markdown 내보내기 함수"""
    exporter = DataExporter(tables_data, sort_config=sort_config)
    return exporter.export_to_markdown(output_path)


def export_to_html(tables_data: Dict, output_path: Path, charts_data: Dict = None, sort_config: Dict = None) -> bool:
    """기존 코드 호환용 HTML 내보내기 함수"""
    exporter = DataExporter(tables_data, charts_data=charts_data, sort_config=sort_config)
    return exporter.export_to_html(output_path)
