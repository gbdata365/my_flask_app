"""
================================================================================
ExportManager - 데이터 내보내기 모듈
================================================================================
Excel, Markdown, HTML 형식으로 데이터를 내보내는 클래스

사용 예시:
    from module.dashboard.export import ExportManager

    # Excel 내보내기
    ExportManager.to_excel(
        data={'시도별': sido_df, '연령별': age_df},
        filepath='output/report.xlsx'
    )

    # Markdown 내보내기
    ExportManager.to_markdown(
        data={'시도별': sido_data, '연령별': age_data},
        filepath='output/report.md',
        title='인구 현황 보고서'
    )

    # 전체 내보내기 (Excel + MD + HTML)
    results = ExportManager.export_all(
        data=data_dict,
        output_dir='output',
        filename='인구현황_202412'
    )
================================================================================
"""
import os
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional, Union
import pandas as pd


class ExportManager:
    """
    데이터 내보내기 관리 클래스

    클래스(Class) 설명:
    ------------------
    이 클래스는 데이터를 다양한 형식(Excel, Markdown, HTML)으로 내보냅니다.
    모든 메서드가 @classmethod이므로 객체 생성 없이 바로 호출할 수 있습니다.

    예시:
        # Excel로 내보내기
        ExportManager.to_excel({'시도별': df}, 'output.xlsx')
    """

    @staticmethod
    def _ensure_dir(filepath: Union[str, Path]) -> Path:
        """디렉토리가 없으면 생성"""
        path = Path(filepath)
        path.parent.mkdir(parents=True, exist_ok=True)
        return path

    @staticmethod
    def _to_dataframe(data: Union[pd.DataFrame, List[Dict], Dict]) -> pd.DataFrame:
        """다양한 형식의 데이터를 DataFrame으로 변환"""
        if isinstance(data, pd.DataFrame):
            return data
        elif isinstance(data, list):
            return pd.DataFrame(data)
        elif isinstance(data, dict) and 'data' in data:
            return pd.DataFrame(data['data'])
        elif isinstance(data, dict):
            return pd.DataFrame([data])
        return pd.DataFrame()

    @classmethod
    def to_excel(
        cls,
        data: Dict[str, Any],
        filepath: Union[str, Path],
        highlight_regions: Optional[List[str]] = None
    ) -> bool:
        """
        Excel 파일로 내보내기 (여러 시트 지원)

        Args:
            data: {'시트명': DataFrame 또는 데이터} 형태의 딕셔너리
            filepath: 저장할 파일 경로
            highlight_regions: 강조할 지역 리스트 (예: ['경상북도'])

        Returns:
            성공 여부
        """
        try:
            filepath = cls._ensure_dir(filepath)

            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                for sheet_name, sheet_data in data.items():
                    df = cls._to_dataframe(sheet_data)
                    if df.empty:
                        continue

                    # 시트명 길이 제한 (Excel 제한: 31자)
                    safe_sheet_name = sheet_name[:31]
                    df.to_excel(writer, sheet_name=safe_sheet_name, index=False)

                    # 스타일 적용 (경상북도 등 강조)
                    if highlight_regions:
                        cls._apply_excel_highlight(
                            writer, safe_sheet_name, df, highlight_regions
                        )

            return True
        except Exception as e:
            print(f"Excel 내보내기 오류: {e}")
            return False

    @classmethod
    def _apply_excel_highlight(
        cls,
        writer: pd.ExcelWriter,
        sheet_name: str,
        df: pd.DataFrame,
        highlight_regions: List[str]
    ):
        """Excel 시트에 강조 스타일 적용"""
        try:
            from openpyxl.styles import PatternFill, Font, Border, Side

            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # 헤더 스타일
            header_fill = PatternFill(start_color='1243A6', end_color='1243A6', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)

            for col_num, col_name in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font

            # 강조 행 스타일 (경상북도 등)
            highlight_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
            highlight_border = Border(
                left=Side(style='medium', color='DC2626'),
                right=Side(style='medium', color='DC2626'),
                top=Side(style='medium', color='DC2626'),
                bottom=Side(style='medium', color='DC2626')
            )

            # 첫 번째 컬럼에서 강조 대상 찾기
            first_col = df.columns[0]
            for row_num, value in enumerate(df[first_col], 2):  # 2부터 (헤더가 1)
                if value in highlight_regions:
                    for col_num in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.fill = highlight_fill
                        cell.border = highlight_border

            # 열 너비 자동 조정
            for col_num, col_name in enumerate(df.columns, 1):
                max_length = max(
                    len(str(col_name)),
                    df[col_name].astype(str).str.len().max() if len(df) > 0 else 0
                )
                worksheet.column_dimensions[chr(64 + col_num)].width = min(max_length + 2, 50)

        except Exception as e:
            print(f"Excel 스타일 적용 오류: {e}")

    @classmethod
    def to_markdown(
        cls,
        data: Dict[str, Any],
        filepath: Union[str, Path],
        title: str = '데이터 보고서',
        highlight_regions: Optional[List[str]] = None
    ) -> bool:
        """
        Markdown 파일로 내보내기

        Args:
            data: {'섹션명': DataFrame 또는 데이터} 형태
            filepath: 저장할 파일 경로
            title: 문서 제목
            highlight_regions: 강조할 지역 리스트

        Returns:
            성공 여부
        """
        try:
            filepath = cls._ensure_dir(filepath)
            highlight_regions = highlight_regions or []

            lines = [
                f'# {title}',
                f'',
                f'생성일시: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}',
                f'',
            ]

            for section_name, section_data in data.items():
                df = cls._to_dataframe(section_data)
                if df.empty:
                    continue

                lines.append(f'## {section_name}')
                lines.append('')

                # 테이블 헤더
                headers = '| ' + ' | '.join(str(col) for col in df.columns) + ' |'
                separator = '| ' + ' | '.join('---' for _ in df.columns) + ' |'
                lines.append(headers)
                lines.append(separator)

                # 테이블 데이터
                for _, row in df.iterrows():
                    row_values = []
                    for val in row:
                        if isinstance(val, float):
                            row_values.append(f'{val:,.2f}')
                        elif isinstance(val, int):
                            row_values.append(f'{val:,}')
                        else:
                            row_values.append(str(val))

                    # 강조 표시 (첫 번째 값이 강조 대상이면 **bold**)
                    if row_values and row_values[0] in highlight_regions:
                        row_values[0] = f'**{row_values[0]}**'

                    lines.append('| ' + ' | '.join(row_values) + ' |')

                lines.append('')

            with open(filepath, 'w', encoding='utf-8') as f:
                f.write('\n'.join(lines))

            return True
        except Exception as e:
            print(f"Markdown 내보내기 오류: {e}")
            return False

    @classmethod
    def to_html(
        cls,
        data: Dict[str, Any],
        filepath: Union[str, Path],
        title: str = '데이터 보고서',
        highlight_regions: Optional[List[str]] = None,
        include_charts: Optional[Dict[str, str]] = None
    ) -> bool:
        """
        HTML 파일로 내보내기 (차트 포함 가능)

        Args:
            data: {'섹션명': DataFrame 또는 데이터} 형태
            filepath: 저장할 파일 경로
            title: 문서 제목
            highlight_regions: 강조할 지역 리스트
            include_charts: {'섹션명': Base64 이미지} 차트 딕셔너리

        Returns:
            성공 여부
        """
        try:
            filepath = cls._ensure_dir(filepath)
            highlight_regions = highlight_regions or []
            include_charts = include_charts or {}

            html = f'''<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <style>
        body {{ font-family: 'Pretendard', 'Noto Sans KR', sans-serif; margin: 20px; background: #f5f5f5; }}
        .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        h1 {{ color: #1243A6; border-bottom: 3px solid #1243A6; padding-bottom: 10px; }}
        h2 {{ color: #1D64F2; margin-top: 30px; }}
        .meta {{ color: #666; font-size: 0.9rem; margin-bottom: 20px; }}
        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; font-size: 0.875rem; }}
        th {{ background: #1243A6; color: white; padding: 10px 8px; text-align: center; }}
        td {{ border: 1px solid #ddd; padding: 8px; text-align: right; }}
        td:first-child {{ text-align: left; font-weight: 500; }}
        tr:nth-child(even) {{ background: #f9fafb; }}
        tr:hover {{ background: #e0f2fe; }}
        tr.highlight {{ background: #fee2e2 !important; border: 2px solid #dc2626; }}
        .chart {{ text-align: center; margin: 20px 0; }}
        .chart img {{ max-width: 100%; height: auto; }}
        @media print {{ body {{ background: white; }} .container {{ box-shadow: none; }} }}
    </style>
</head>
<body>
    <div class="container">
        <h1>{title}</h1>
        <p class="meta">생성일시: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
'''

            for section_name, section_data in data.items():
                df = cls._to_dataframe(section_data)
                if df.empty:
                    continue

                html += f'<h2>{section_name}</h2>\n'

                # 차트가 있으면 추가
                if section_name in include_charts:
                    chart_img = include_charts[section_name]
                    html += f'<div class="chart"><img src="data:image/png;base64,{chart_img}" alt="{section_name} 차트"></div>\n'

                # 테이블
                html += '<table>\n<thead><tr>'
                for col in df.columns:
                    html += f'<th>{col}</th>'
                html += '</tr></thead>\n<tbody>\n'

                for _, row in df.iterrows():
                    # 강조 대상 확인
                    is_highlight = str(row.iloc[0]) in highlight_regions
                    row_class = ' class="highlight"' if is_highlight else ''

                    html += f'<tr{row_class}>'
                    for val in row:
                        if isinstance(val, float):
                            html += f'<td>{val:,.2f}</td>'
                        elif isinstance(val, int):
                            html += f'<td>{val:,}</td>'
                        else:
                            html += f'<td>{val}</td>'
                    html += '</tr>\n'

                html += '</tbody></table>\n'

            html += '''
    </div>
</body>
</html>'''

            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(html)

            return True
        except Exception as e:
            print(f"HTML 내보내기 오류: {e}")
            return False

    @classmethod
    def export_all(
        cls,
        data: Dict[str, Any],
        output_dir: Union[str, Path],
        filename: str,
        title: str = '데이터 보고서',
        highlight_regions: Optional[List[str]] = None,
        charts: Optional[Dict[str, str]] = None,
        formats: List[str] = None
    ) -> Dict[str, Any]:
        """
        여러 형식으로 한 번에 내보내기

        Args:
            data: {'섹션명': 데이터} 딕셔너리
            output_dir: 출력 디렉토리
            filename: 파일명 (확장자 제외)
            title: 문서 제목
            highlight_regions: 강조할 지역 리스트
            charts: 차트 이미지 딕셔너리
            formats: 내보낼 형식 리스트 ['xlsx', 'md', 'html'] (기본: 전부)

        Returns:
            {'success': True/False, 'files': [생성된 파일 경로들], 'errors': [오류 목록]}
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        formats = formats or ['xlsx', 'md', 'html']
        highlight_regions = highlight_regions or ['경상북도']

        results = {
            'success': True,
            'files': [],
            'errors': [],
            'output_dir': str(output_dir)
        }

        # Excel
        if 'xlsx' in formats:
            xlsx_path = output_dir / f'{filename}.xlsx'
            if cls.to_excel(data, xlsx_path, highlight_regions):
                results['files'].append(str(xlsx_path))
            else:
                results['errors'].append('Excel 내보내기 실패')

        # Markdown
        if 'md' in formats:
            md_path = output_dir / f'{filename}.md'
            if cls.to_markdown(data, md_path, title, highlight_regions):
                results['files'].append(str(md_path))
            else:
                results['errors'].append('Markdown 내보내기 실패')

        # HTML
        if 'html' in formats:
            html_path = output_dir / f'{filename}.html'
            if cls.to_html(data, html_path, title, highlight_regions, charts):
                results['files'].append(str(html_path))
            else:
                results['errors'].append('HTML 내보내기 실패')

        results['success'] = len(results['errors']) == 0
        return results
